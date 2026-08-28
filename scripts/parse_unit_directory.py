#!/usr/bin/env python3
"""Parse the Yardi Unit Directory export.

The directory is the building's fixed description: every unit, its floorplan
code, its square footage, and how many bedrooms and baths that plan has. It
carries no lease and no resident — it is what a unit *is*, not who is in it —
so nothing here needs scrubbing, and the ``Notes`` column is dropped anyway
rather than trusted to stay empty.

One export covers several properties. Each begins with a bare property-code row
(``p0005611``) and ends with a ``Total p0005611`` row carrying the property's
own unit count, rent total and square footage. Those totals are the tie-out:
the rows this parser keeps must reproduce all three exactly, or the section is
refused rather than published. A final ``Grand Total`` row is checked the same
way against the sections together.

Two unit ids at The Landing are placeholders, not apartments — ``WAITLIST`` and
``WAIT1B1B``, which is why the directory says 265 units where the rent roll
says 263. They are counted separately as ``placeholder_units`` rather than
dropped silently, since the export's own total includes them and the tie-out
has to as well.

Emits, per section: the floorplan table (bedrooms, baths, unit count, square
footage range) and the property's counts. Aggregates only; the per-unit rows
are read for the tie-out and to build the plans, and are not returned.

  python scripts/parse_unit_directory.py <UnitDirectory.xlsx>
"""
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from xlsx_anchors import LayoutError, norm  # noqa: E402

REPORT_TYPE = "unit_directory"

# The header row, by label. Anchored rather than positional: the export has
# grown a column before (Notes) and would otherwise have shifted Baths into it.
COLUMNS = {
    "unit": ("unit",),
    "address": ("address",),
    "plan": ("unit type",),
    "rent": ("rent",),
    "deposit": ("deposit",),
    "sqft": ("sqft", "square feet"),
    "beds": ("room", "rooms", "bedrooms"),
    "baths": ("baths", "bath"),
}
CODE = re.compile(r"^p\d{6,7}$|^rs\w+$|^cam\w+$|^dnc\w+$|^esx\d+$|^lm\d{5}$"
                  r"|^bec\d{4}$|^wcc\d{4}$", re.I)
TOTAL = re.compile(r"^total\s+(\S+)$", re.I)
UNITS_IN_TOTAL = re.compile(r"units:\s*([\d,]+)", re.I)
# A unit id that is not an apartment: the waitlist placeholders Yardi carries
# so a prospect can be attached to a property with no unit yet.
PLACEHOLDER = re.compile(r"^wait", re.I)
CENT = 0.01


def _num(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[,$\s]", "", str(v))
    if s in ("", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _header(ws):
    """{field: column} for the directory's header row, and the row it is on."""
    for r in range(1, min(ws.max_row, 15) + 1):
        got = {}
        for c in range(1, ws.max_column + 1):
            label = norm(ws.cell(row=r, column=c).value)
            for field, names in COLUMNS.items():
                if label in names and field not in got:
                    got[field] = c
        if {"unit", "plan", "sqft", "beds"} <= set(got):
            return got, r
    raise LayoutError("no Unit Directory header row found (need Unit, Unit "
                      "Type, Sqft and Room columns in the first 15 rows)")


def _as_of(path):
    """The export's date, from its filename: UnitDirectory08_25_2026.xlsx."""
    m = re.search(r"(\d{2})[_-](\d{2})[_-](\d{4})", os.path.basename(path))
    if m:
        return f"{m.group(3)}-{m.group(1)}-{m.group(2)}"
    m = re.search(r"(\d{4})[_-]?(\d{2})[_-]?(\d{2})", os.path.basename(path))
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else None


def _plans(rows):
    """Floorplan table: one entry per unit-type code seen in the section.

    Bedrooms and baths belong to the plan, not the unit, so a plan whose rows
    disagree is reported rather than averaged into a number that describes no
    apartment.
    """
    out = {}
    for r in rows:
        p = out.setdefault(r["plan"], {"units": 0, "sqft": [], "rents": [],
                                       "beds": set(), "baths": set()})
        p["units"] += 1
        if r["sqft"] is not None:
            p["sqft"].append(r["sqft"])
        if r["rent"] is not None:
            p["rents"].append(r["rent"])
        if r["beds"] is not None:
            p["beds"].add(int(r["beds"]))
        if r["baths"] is not None:
            p["baths"].add(round(r["baths"], 2))
    plans, mixed = {}, []
    for code, p in sorted(out.items()):
        if len(p["beds"]) > 1:
            mixed.append(f"{code}: bedrooms {sorted(p['beds'])}")
        if len(p["baths"]) > 1:
            mixed.append(f"{code}: baths {sorted(p['baths'])}")
        plans[code] = {
            "bedrooms": min(p["beds"]) if p["beds"] else None,
            "baths": min(p["baths"]) if p["baths"] else None,
            "units": p["units"],
            "sqft_min": min(p["sqft"]) if p["sqft"] else None,
            "sqft_max": max(p["sqft"]) if p["sqft"] else None,
            "sqft_avg": round(sum(p["sqft"]) / len(p["sqft"])) if p["sqft"] else None,
            "rent_min": min(p["rents"]) if p["rents"] else None,
            "rent_max": max(p["rents"]) if p["rents"] else None,
        }
    return plans, mixed


def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Report1"] if "Report1" in wb.sheetnames else wb.worksheets[0]
    fields, hdr = _header(ws)

    sections, cur, problems = [], None, []
    grand = None

    def close(declared):
        """Finish the open section against the Total row that ends it."""
        if cur is None:
            return
        real = [r for r in cur["rows"] if not PLACEHOLDER.match(r["unit"])]
        holders = [r for r in cur["rows"] if PLACEHOLDER.match(r["unit"])]
        plans, mixed = _plans(cur["rows"])
        sqft = sum(r["sqft"] or 0 for r in cur["rows"])
        rent = sum(r["rent"] or 0 for r in cur["rows"])
        checks = []
        # The export's own Total row is the only independent statement of what
        # the section should contain, so all three of its numbers are checked.
        if declared:
            for what, got, want in (("unit count", len(cur["rows"]), declared["units"]),
                                    ("rent total", rent, declared["rent"]),
                                    ("sqft total", sqft, declared["sqft"])):
                if want is None:
                    continue
                gap = abs(got - want)
                checks.append({"check": what, "parsed": got, "report": want,
                               "gap": round(gap, 2)})
                if gap > CENT:
                    problems.append(
                        f"{cur['property_code']}: {what} does not tie out -- "
                        f"parsed {got:,.2f} vs the report's {want:,.2f}")
        else:
            problems.append(f"{cur['property_code']}: no 'Total' row to tie out against")
        problems.extend(f"{cur['property_code']}: {m}" for m in mixed)
        sections.append({
            "property_code": cur["property_code"],
            "units": len(cur["rows"]),
            "residential_units": len(real),
            "placeholder_units": len(holders),
            "placeholders": sorted(r["unit"] for r in holders),
            "sqft_total": round(sqft, 2),
            "rent_total": round(rent, 2),
            "plans": plans,
            "checks": checks,
        })

    for r in range(hdr + 1, ws.max_row + 1):
        a = ws.cell(row=r, column=fields["unit"]).value
        label = str(a).strip() if a is not None else ""
        if not label:
            continue
        m = TOTAL.match(label)
        if m or norm(label) == "grand total":
            um = UNITS_IN_TOTAL.search(
                " ".join(str(ws.cell(row=r, column=c).value or "")
                         for c in range(1, ws.max_column + 1)))
            declared = {
                "units": int(um.group(1).replace(",", "")) if um else None,
                "rent": _num(ws.cell(row=r, column=fields["rent"]).value),
                "sqft": _num(ws.cell(row=r, column=fields["sqft"]).value),
            }
            if m:
                close(declared)
                cur = None
            else:
                grand = declared
            continue
        if CODE.match(label) and not ws.cell(row=r, column=fields["plan"]).value:
            close(None) if cur is not None else None
            cur = {"property_code": label, "rows": []}
            continue
        if cur is None:
            continue
        plan = ws.cell(row=r, column=fields["plan"]).value
        if not plan:
            continue
        cur["rows"].append({
            "unit": label,
            "plan": str(plan).strip(),
            "rent": _num(ws.cell(row=r, column=fields["rent"]).value),
            "sqft": _num(ws.cell(row=r, column=fields["sqft"]).value),
            "beds": _num(ws.cell(row=r, column=fields.get("beds", 0)).value)
                    if fields.get("beds") else None,
            "baths": _num(ws.cell(row=r, column=fields.get("baths", 0)).value)
                     if fields.get("baths") else None,
        })
    close(None) if cur is not None else None

    checks = []
    if grand and grand["units"] is not None:
        got = sum(s["units"] for s in sections)
        checks.append({"check": "grand total unit count", "parsed": got,
                       "report": grand["units"], "gap": abs(got - grand["units"])})
        if got != grand["units"]:
            problems.append(f"grand total unit count does not tie out -- sections "
                            f"sum to {got}, the report says {grand['units']}")

    return {
        "report_type": REPORT_TYPE,
        "source_file": os.path.basename(path),
        "as_of": _as_of(path),
        "sections": sections,
        "checks": checks,
        "problems": problems,
    }


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    import json
    p = parse(sys.argv[1])
    for s in p["sections"]:
        print(f"{s['property_code']}: {s['units']} units "
              f"({s['residential_units']} residential + {s['placeholder_units']} "
              f"placeholder), {len(s['plans'])} plans, "
              f"{s['sqft_total']:,.0f} sqft, ${s['rent_total']:,.0f} rent")
        for c in s["checks"]:
            print(f"    {c['check']}: parsed {c['parsed']:,.2f} vs report "
                  f"{c['report']:,.2f} (gap {c['gap']})")
    for c in p["checks"]:
        print(f"{c['check']}: {c['parsed']} vs {c['report']}")
    for pr in p["problems"]:
        print(f"PROBLEM: {pr}")
    print(json.dumps({s["property_code"]: s["plans"] for s in p["sections"]},
                     indent=1)[:400])
    sys.exit(1 if p["problems"] else 0)
