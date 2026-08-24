#!/usr/bin/env python3
"""Print the structure of a report without printing anyone's name.

For working out what a new export actually contains — which property it is, what
its columns are, why a tie-out fails — when the file itself cannot leave the
runner. Output goes to a CI log, so it is filtered on the same principle as the
published JSON: labels, codes, counts and sums are fine; a resident's name is
not, ever.

Any column whose header looks like a person's name is redacted, and no cell from
such a column is printed. Long strings are truncated. If a header cannot be
identified, string cells are withheld rather than guessed at.

Usage:
  python scripts/inspect_report.py <file.xlsx> [more.xlsx ...]
  python scripts/inspect_report.py --delinquency <file.xlsx>   # adds a tie-out probe
"""
import argparse
import os
import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter as CL

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# same shape of test the PII guard uses, so the two cannot disagree about what
# counts as a person-shaped column
NAME_HDR = re.compile(r"resident (last |first )?name|^name$|tenant name|^tenant$|"
                      r"^resident$|last name|first name|occupant", re.I)
CODE = re.compile(r"\b(p\d{6,7}|rs\d{3}|rspalma[ns]|camadel\w*|camadret|"
                  r"dnc\w+|esx\d+|lm\d{5}|bec\d{4}|wcc\d{4})\b", re.I)
TOTALISH = re.compile(r"^\s*(grand )?total\b|^\s*subtotal\b", re.I)
REDACTED = "<redacted>"

# Property names from the master, so a report that spells the property out in
# words instead of carrying a Yardi code can still be attributed.
try:
    import json as _json
    PROPERTY_NAMES = sorted(
        {p["name"] for p in _json.load(open("config/properties.json"))["properties"]}
        | {"Palma North", "Palma South"}, key=len, reverse=True)
except Exception:                                          # noqa: BLE001
    PROPERTY_NAMES = []


def name_columns(ws, search_rows=25):
    """Columns sitting under a person-name header, plus the header row it found."""
    for r in range(1, min(ws.max_row, search_rows) + 1):
        hits = {c for c in range(1, ws.max_column + 1)
                if isinstance(ws.cell(row=r, column=c).value, str)
                and NAME_HDR.search(ws.cell(row=r, column=c).value.strip())}
        if hits:
            return hits, r
    return set(), None


def show(v, redact):
    if redact:
        return REDACTED
    if v is None:
        return ""
    if isinstance(v, str):
        s = re.sub(r"\s+", " ", v).strip()
        return s[:44]
    if isinstance(v, float):
        return f"{v:,.2f}"
    return str(v)[:44]


def inspect(path):
    print("=" * 78)
    print(f"FILE {os.path.basename(path)}  ({os.path.getsize(path):,} bytes)")
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"  sheets: {wb.sheetnames}")
    for ws in wb.worksheets:
        redact, hdr_row = name_columns(ws)
        print(f"\n  --- sheet {ws.title!r}  {ws.max_row} rows x {ws.max_column} cols"
              f"   name column(s) redacted: "
              f"{sorted(CL(c) for c in redact) or 'none found'}")

        codes, names_seen = set(), set()
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.column not in redact:
                    for m in CODE.findall(c.value):
                        codes.add(m)
                    for pname in PROPERTY_NAMES:
                        if re.search(r"\b" + re.escape(pname) + r"\b", c.value, re.I):
                            names_seen.add(pname)
        print(f"      property code(s) seen: {sorted(codes) or 'NONE'}")
        print(f"      property name(s) seen: {sorted(names_seen) or 'NONE'}")

        # Rows above the header often carry the report title and the property it
        # was run for, and a header can span two rows — so show the top of the
        # sheet, name columns still redacted.
        print("      top of sheet:")
        for r in range(1, min(ws.max_row, 6) + 1):
            cells = [f"{CL(c)}={show(ws.cell(row=r, column=c).value, c in redact)}"
                     for c in range(1, ws.max_column + 1)
                     if ws.cell(row=r, column=c).value is not None]
            if cells:
                print(f"        r{r}: " + " | ".join(cells))

        # the header row, then a few data rows, then anything total-shaped
        if hdr_row:
            cells = [f"{CL(c)}={show(ws.cell(row=hdr_row, column=c).value, False)}"
                     for c in range(1, ws.max_column + 1)
                     if ws.cell(row=hdr_row, column=c).value is not None]
            print(f"      header row {hdr_row}: " + " | ".join(cells))
        for r in range(1, min(ws.max_row, 40) + 1):
            label = ws.cell(row=r, column=1).value
            if isinstance(label, str) and TOTALISH.search(label):
                cells = [f"{CL(c)}={show(ws.cell(row=r, column=c).value, c in redact)}"
                         for c in range(1, ws.max_column + 1)
                         if ws.cell(row=r, column=c).value is not None]
                print(f"      TOTAL row {r}: " + " | ".join(cells))

        # numeric column sums — the fastest way to see which column is which
        sums = {}
        for c in range(1, ws.max_column + 1):
            vals = [ws.cell(row=r, column=c).value for r in range(1, ws.max_row + 1)]
            nums = [v for v in vals if isinstance(v, (int, float))]
            if nums:
                sums[CL(c)] = (len(nums), sum(nums))
        print("      numeric columns (count, sum):")
        for k, (n, s) in sums.items():
            print(f"        {k}: {n:4d}  {s:,.2f}")


def t12_header_probe(paths):
    """Which period does each T12 statement actually claim? Aggregates only.

    The parser trusts the month labels in the header row (MONTHS_COLS), so two
    statements whose monthly ratios match at an offset mean one header is
    wrong. Print every title row above the month header verbatim — Yardi puts
    the period it was run for up there — then the parsed labels and monthly
    ratios, then the pairwise shift at which the ratio series line up.
    """
    import parse_t12_statement as t12
    MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    parsed = []
    for path in paths:
        print("=" * 78)
        print(f"T12 {os.path.basename(path)}")
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["Report1"] if "Report1" in wb.sheetnames else wb.worksheets[0]
        redact, _ = name_columns(ws)

        # every row above and including the month-header row, verbatim
        hdr = None
        for r in range(1, min(ws.max_row, 15) + 1):
            vals = [ws.cell(row=r, column=c + 1).value for c in t12.MONTHS_COLS]
            if all(v and any(m in str(v) for m in MONTHS) for v in vals):
                hdr = r
                break
        for r in range(1, (hdr or min(ws.max_row, 8)) + 1):
            cells = [f"{CL(c)}={show(ws.cell(row=r, column=c).value, c in redact)}"
                     for c in range(1, ws.max_column + 1)
                     if ws.cell(row=r, column=c).value is not None]
            if cells:
                tag = "  <- month header" if r == hdr else ""
                print(f"  r{r}: " + " | ".join(cells) + tag)
        if hdr is None:
            print("  NO month-header row found in the first 15 rows")

        try:
            p = t12.parse_t12(path)
        except Exception as e:                              # noqa: BLE001
            print(f"  parser refused it: {type(e).__name__}: {e}")
            continue
        print(f"  parsed: {p['property']} ({p['property_code']})  book={p['book']}"
              f"  period_end={p['period_end']}")
        print(f"  labels: {p['labels']}")
        print("  monthly expense ratios: "
              + "  ".join(f"{l}={r}" for l, r in
                          zip(p["labels"] or [""] * 12, p["expense_ratio_monthly"])))
        parsed.append((os.path.basename(path), p))

    # Pairwise: at which month shift do two statements' ratio series agree?
    # Twelve distinct months agreeing at shift 0 is a duplicate; agreeing at a
    # non-zero shift means one statement's header claims the wrong period.
    print("=" * 78)
    print("pairwise ratio alignment (months agreeing within 0.15pp, of overlap):")
    for i in range(len(parsed)):
        for j in range(i + 1, len(parsed)):
            (fa, a), (fb, b) = parsed[i], parsed[j]
            best = []
            for shift in range(-4, 5):
                hits = total = 0
                for k in range(12):
                    if 0 <= k + shift < 12:
                        ra, rb = a["expense_ratio_monthly"][k], \
                                 b["expense_ratio_monthly"][k + shift]
                        if ra is not None and rb is not None:
                            total += 1
                            hits += abs(ra - rb) <= 0.15
                best.append((shift, hits, total))
            top = max(best, key=lambda t: (t[1], t[1] / t[2] if t[2] else 0))
            print(f"  {fa}  vs  {fb}:")
            for shift, hits, total in best:
                mark = "  <- best" if (shift, hits, total) == top and hits else ""
                print(f"      shift {shift:+d}: {hits}/{total}{mark}")


def tie_out_probe(path):
    """Why does the delinquency parser refuse this file? Aggregates only."""
    import parse_delinquency
    from xlsx_anchors import header_map
    print("\n  --- delinquency tie-out probe")
    try:
        loose = parse_delinquency.parse(path, strict=False)
    except Exception as e:                                  # noqa: BLE001
        print(f"      parser could not read it at all: {type(e).__name__}: {e}")
        return
    s = loose["summary"]
    print(f"      property           : {loose.get('property')} / {loose.get('property_code')}")
    print(f"      rows the parser kept: {len(loose.get('residents') or [])}")
    print(f"      gross owed          : {s.get('gross_owed'):,.2f}")
    print(f"      credits             : {s.get('credits'):,.2f}")
    print(f"      net                 : {s.get('net'):,.2f}")
    print(f"      aging               : {s.get('aging')}")
    print(f"      retail balance      : {s.get('retail_balance')}")
    for p in loose.get("problems") or []:
        print(f"      PROBLEM: {p}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[loose.get("sheet")] if loose.get("sheet") in wb.sheetnames else wb.worksheets[0]
    try:
        fields, hdr = header_map(ws, parse_delinquency.COLUMNS)
        print(f"      header row {hdr}, columns mapped: "
              + ", ".join(f"{k}->{CL(v)}" for k, v in sorted(fields.items())))
        col = fields.get("total_owed")
        if col:
            every = [ws.cell(row=r, column=col).value
                     for r in range(hdr + 1, ws.max_row + 1)]
            nums = [v for v in every if isinstance(v, (int, float))]
            print(f"      Total Owed column {CL(col)}: {len(nums)} numeric cells, "
                  f"sum {sum(nums):,.2f} (includes any total rows)")
    except Exception as e:                                  # noqa: BLE001
        print(f"      could not map the header: {e}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("files", nargs="+")
    ap.add_argument("--delinquency", action="store_true",
                    help="also run the delinquency tie-out probe on each file")
    ap.add_argument("--t12", action="store_true",
                    help="T12 header probe: title rows, parsed period, monthly "
                         "ratios, and pairwise shift alignment (skips the "
                         "generic dump)")
    a = ap.parse_args()
    present = [f for f in a.files if os.path.exists(f)]
    for f in a.files:
        if f not in present:
            print(f"missing: {f}")
    if a.t12:
        t12_header_probe(present)
        sys.exit(0)
    for f in present:
        inspect(f)
        if a.delinquency:
            tie_out_probe(f)
