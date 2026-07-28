#!/usr/bin/env python3
"""Prove the extractor's guards fire, rather than only that it works on a good file.

Two halves:

1. Anchor unit tests on synthetic sheets — a month added, a block that grew, a
   label that moved. These are the changes a new period actually brings, and
   the old fixed-coordinate reader got every one of them wrong.
2. End-to-end negative tests against the real workbook, mutated to break in the
   ways it plausibly will. Each must exit non-zero, because silently publishing
   a wrong number is the failure worth preventing.

Usage: python scripts/test_extract_landing.py <workbook.xlsx>
"""
import os
import shutil
import subprocess
import sys
import tempfile

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from xlsx_anchors import (LayoutError, block, find_row, month_axis,  # noqa: E402
                          parse_month)

HERE = os.path.dirname(os.path.abspath(__file__))
EXTRACT = os.path.join(HERE, "extract_landing.py")
PASS, FAIL = [], []


def ok(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    print(f"  {'PASS' if cond else 'FAIL'}  {name}" + (f"  — {detail}" if detail else ""))


def sheet(rows):
    """Build a throwaway worksheet from a list of row lists (1-indexed cols)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T"
    for r, row in enumerate(rows, start=1):
        for c, v in enumerate(row, start=1):
            if v is not None:
                ws.cell(row=r, column=c, value=v)
    return ws


# ---------------------------------------------------------------- unit tests
print("\nanchor unit tests")

ws = sheet([
    [None, "Line item", "Jan 2025", "Feb 2025", "Mar 2025", "TTM (x)"],
    [None, "Revenue", 1, 2, 3, 6],
])
months, cols, ttm = month_axis(ws, 1)
ok("month axis reads header", months == ["2025-01", "2025-02", "2025-03"], str(months))
ok("TTM column found by label", ttm == 6, f"col {ttm}")

# The case that broke the old reader: one more month shifts TTM one column right.
ws = sheet([
    [None, "Line item", "Jan 2025", "Feb 2025", "Mar 2025", "Apr 2025", "TTM (x)"],
    [None, "Revenue", 1, 2, 3, 4, 10],
])
months, cols, ttm = month_axis(ws, 1)
ok("added month extends axis", months[-1] == "2025-04" and len(months) == 4, str(months))
ok("added month moves TTM with it", ttm == 7, f"col {ttm} (was 6 with 3 months)")

ws = sheet([
    [None, "Line item", "Jan 2025", "Mar 2025", "TTM"],
])
try:
    month_axis(ws, 1)
    ok("gap in months raises", False, "no exception")
except LayoutError as e:
    ok("gap in months raises", "consecutive" in str(e), str(e)[:60])

ws = sheet([[None, "Line item", "Jan 2025", "Feb 2025"]])
try:
    month_axis(ws, 1)
    ok("missing TTM column raises", False, "no exception")
except LayoutError as e:
    ok("missing TTM column raises", "TTM" in str(e) or "ttm" in str(e), str(e)[:60])

# A block that grows must be read whole, and must stop at its Total row.
ws = sheet([
    [None, "Rank", "Unit", "Gap"],
    [None, 1, "355", 45276],
    [None, 2, "664", 43680],
    [None, 3, "257", 43524],
    [None, "Total", None, 132480],
    [None, "some footnote"],
])
rows = block(ws, 1, cols=[2, 3, 4])
ok("block reads to the Total sentinel", len(rows) == 3, f"{len(rows)} rows")
ok("block excludes the Total row", all(r[0] != "Total" for r in rows))

ws = sheet([
    [None, "Rank", "Unit", "Gap"],
    [None, 1, "355", 45276],
    [None, 2, "664", 43680],
    [None, 3, "257", 43524],
    [None, 4, "221", 43476],
    [None, "Total", None, 175956],
])
ok("block grows with the data", len(block(ws, 1, cols=[2, 3, 4])) == 4)

# keep() must drop summary lines that share the label column with real rows.
ws = sheet([
    [None, "Unit", "Type", "Sq ft", "Market"],
    [None, "101", "laa1", 560, 5159],
    [None, "102", "laa5", 688, 5434],
    [None, "TOTALS / SUMMARY"],
    [None, "Units", 263],
])
rows = block(ws, 1, cols=[2, 3, 4, 5], stop_on_blank=False,
             keep=lambda v: isinstance(v[2], (int, float)) and isinstance(v[3], (int, float)))
ok("keep() filters summary rows", len(rows) == 2, f"{len(rows)} unit rows kept")

# Labels that repeat in different sections must be separable by scope.
ws = sheet([
    [None, "Line item", "Jan 2025"],
    [None, "Vacancy loss", -100],
    [None, "AS % OF POTENTIAL"],
    [None, "Vacancy loss", -0.05],
])
top = find_row(ws, "Vacancy loss", before=find_row(ws, "AS % OF POTENTIAL"))
bot = find_row(ws, "Vacancy loss", after=find_row(ws, "AS % OF POTENTIAL"))
ok("repeated label separable by scope", (top, bot) == (2, 4), f"rows {top} and {bot}")

ok("month parser handles all seen forms",
   [parse_month(x) for x in ["Jan 2025", "2026-08", "Sep 2025", "2026-07-14"]]
   == ["2025-01", "2026-08", "2025-09", "2026-07"])
ok("month parser rejects non-months",
   [parse_month(x) for x in ["TTM (Aug25-Jul26)", "Total", None, "GL"]] == [None] * 4)

# --------------------------------------------------------- end-to-end tests
if len(sys.argv) > 1:
    src = sys.argv[1]
    print("\nend-to-end negative tests")
    tmp = tempfile.mkdtemp()

    def run(path, out):
        r = subprocess.run([sys.executable, EXTRACT, path, "--out", out],
                           capture_output=True, text=True)
        return r.returncode, (r.stdout + r.stderr)

    good = os.path.join(tmp, "good.xlsx")
    shutil.copy(src, good)
    code, log = run(good, os.path.join(tmp, "good.json"))
    ok("unmodified workbook extracts cleanly", code == 0, f"exit {code}")

    # 1. saved without recalculating -> every formula reads None
    stale = os.path.join(tmp, "stale.xlsx")
    shutil.copy(src, stale)
    wb = openpyxl.load_workbook(stale)
    wb.save(stale)                      # openpyxl drops cached values on save
    code, log = run(stale, os.path.join(tmp, "stale.json"))
    ok("stale workbook is refused", code != 0 and "recalculat" in log,
       f"exit {code}")
    ok("stale workbook writes no output",
       not os.path.exists(os.path.join(tmp, "stale.json")))

    # 2. an anchor label renamed -> must raise, not guess
    renamed = os.path.join(tmp, "renamed.xlsx")
    shutil.copy(src, renamed)
    wb = openpyxl.load_workbook(renamed)
    r = find_row(wb["Inputs"], "Total units")
    wb["Inputs"].cell(row=r, column=2, value="Unit count")
    wb.save(renamed)
    code, log = run(renamed, os.path.join(tmp, "renamed.json"))
    ok("renamed anchor label is refused", code != 0, f"exit {code}")

    # 3. a unit row deleted -> count no longer matches Inputs
    short = os.path.join(tmp, "short.xlsx")
    shutil.copy(src, short)
    wb = openpyxl.load_workbook(short, data_only=True)   # values only: keeps them readable
    ug = wb["Unit Gap Analysis"]
    hdr = find_row(ug, "Unit")
    ug.delete_rows(hdr + 1)
    wb.save(short)
    code, log = run(short, os.path.join(tmp, "short.json"))
    ok("dropped unit row is caught", code != 0 and "unit count" in log.lower(),
       f"exit {code}")

    # 4. a tie-out row broken -> statement no longer reconciles
    tie = os.path.join(tmp, "tie.xlsx")
    shutil.copy(src, tie)
    wb = openpyxl.load_workbook(tie, data_only=True)
    rc = wb["Rent Capture"]
    r = find_row(rc, "Check vs statement (should be nil)")
    rc.cell(row=r, column=4, value=12345)
    wb.save(tie)
    code, log = run(tie, os.path.join(tmp, "tie.json"))
    ok("broken tie-out is caught", code != 0 and "tie-out" in log.lower(), f"exit {code}")

    shutil.rmtree(tmp, ignore_errors=True)
else:
    print("\n(pass a workbook path to also run the end-to-end negative tests)")

print(f"\n{len(PASS)} passed, {len(FAIL)} failed")
if FAIL:
    for n in FAIL:
        print("  failed: " + n)
sys.exit(1 if FAIL else 0)
