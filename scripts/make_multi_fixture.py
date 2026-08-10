"""Build a two-property delinquency fixture matching the real Palma export.

Structure, columns and every reported total come from what the inspector read out
of '2026-08-10 Delinquency_8_1_2026.xls.xlsx': two sections (rspalman, rspalmas),
8 + 1 data rows, section totals, then one Grand Total. Per-unit values are
invented but constructed so each section reproduces the real figures exactly —
including the shape that matters here, charges and prepayments sitting on
SEPARATE rows, which is why Palma nets to a credit while its aging is positive.
"""
import openpyxl, sys

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Report1"
ws["A1"] = "Delinquency Summary Report"
ws["A2"] = "Delinquency Summary as of 08/01/2026"
h1 = ["Property", "Resident", "Resident", "Resident", "Total", "Future",
      "0-30", "31-60", "61-90", "Over 90", "Prepayments", "Total"]
h2 = ["Unit", "Code", "Last Name", "Status", "Charges", "Charges",
      "Owed", "Owed", "Owed", "Owed", "", "Owed"]
for i, (a, b) in enumerate(zip(h1, h2), start=1):
    ws.cell(row=3, column=i, value=a)
    ws.cell(row=4, column=i, value=b)

def row(r, unit, code, name, d30=0.0, d60=0.0, d90=0.0, o90=0.0, prepay=0.0):
    charges = round(d30 + d60 + d90 + o90, 2)
    owed = round(charges + prepay, 2)
    for c, v in enumerate([unit, code, name, "Current", charges, 0,
                           d30, d60, d90, o90, prepay, owed], start=1):
        ws.cell(row=r, column=c, value=v)

# --- rspalman: 5 charge rows + 3 prepayment rows = 8, as the real file has
ws["A5"] = "rspalman - Palma North"
row(6,  "101", "t1001", "Aaa", d30=10000.00)
row(7,  "102", "t1002", "Bbb", d30=4989.88)
row(8,  "103", "t1003", "Ccc", d60=259.29)
row(9,  "104", "t1004", "Ddd", d90=10345.54)
row(10, "105", "t1005", "Eee", o90=11846.64)
row(11, "106", "t1006", "Fff", prepay=-20000.00)
row(12, "107", "t1007", "Ggg", prepay=-20000.00)
row(13, "108", "t1008", "Hhh", prepay=-17927.71)
for c, v in enumerate(["Total rspalman - Palma North", None, None, None,
                       37441.35, 0, 14989.88, 259.29, 10345.54, 11846.64,
                       -57927.71, -20486.36], start=1):
    ws.cell(row=14, column=c, value=v)

# --- rspalmas: one credit row
ws["A15"] = "rspalmas - Palma South"
row(16, "201", "t2001", "Iii", d30=-74.00, d90=-76.00)
for c, v in enumerate(["Total rspalmas - Palma South", None, None, None,
                       -150, 0, -74, 0, -76, 0, 0, -150], start=1):
    ws.cell(row=17, column=c, value=v)

for c, v in enumerate(["Grand Total", None, None, None,
                       37291.35, 0, 14915.88, 259.29, 10269.54, 11846.64,
                       -57927.71, -20636.36], start=1):
    ws.cell(row=18, column=c, value=v)

wb.save(sys.argv[1]); print("wrote", sys.argv[1])
