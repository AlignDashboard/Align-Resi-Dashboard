#!/usr/bin/env python3
"""Build docs/landing.json — data for The Landing's property view.

Reads the analyst workbook's computed output tabs. Those tabs are fed by the
grey "Source *" tabs, which are pasted from the Yardi/RealPage reports listed on
the workbook's own Data Lineage tab:

    Source CY25 / Aug25-Jul26  <- 12_Month_Statement_<code>_Accrual  (T12 Expenses)
    Source Rent Roll Jul / Jun <- SPV PM Deliverable Package, Rent Roll tab
    Source Delinquency         <- rs_rp_DelinquencySummaryReport
    Source Renewal Tracker     <- Landing 2025 Renewal Tracker (monthly sheets + MTM tab)
    Lease Detail               <- RealPage rental rate tracker (TYPED IN, not a grey tab)

To refresh: paste the new reports into the grey tabs, let Excel recalculate,
save, then run this script. It re-derives everything, so nothing is carried
over from the previous run.

Everything is located by label rather than by cell coordinate (see
xlsx_anchors.py), because row and column counts move every period — a month is
added, a lease is signed, a holdover is repriced. Validation runs at the end and
exits non-zero on anything that would put wrong numbers on the dashboard.

Usage:
  python scripts/extract_landing.py <workbook.xlsx> [--out docs/landing.json] [--allow-warnings]
"""
import argparse
import datetime
import json
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from xlsx_anchors import (LayoutError, block, cell, find_row, is_stale,  # noqa: E402
                          labelled_series, month_axis, norm, parse_month, scalar,
                          series)

ap = argparse.ArgumentParser()
ap.add_argument("workbook")
ap.add_argument("--out", default="docs/landing.json")
ap.add_argument("--allow-warnings", action="store_true",
                help="exit 0 even if soft checks fail (fatal checks still fail)")
args = ap.parse_args()

wb = openpyxl.load_workbook(args.workbook, data_only=True)

FATAL, WARN = [], []
def fatal(msg): FATAL.append(msg)
def warn(msg): WARN.append(msg)
def check(ok, msg, hard=True):
    if not ok:
        (fatal if hard else warn)(msg)
    return ok

# V37 renamed the Holdovers tab to MTM and added MTM Analysis (tracker
# reconciliation), Scorecard (scored insights) and a Source Renewal Tracker tab.
REQUIRED_SHEETS = ["Inputs", "Rent Capture", "Expense & NOI", "Expense Overview",
                   "Delinquency", "MTM", "MTM Analysis", "Renewal Pipeline",
                   "Unit Gap Analysis", "Rate vs Occupancy", "Lease Detail",
                   "Floorplan & Rollover", "Scorecard"]
missing_sheets = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
if missing_sheets:
    print("FATAL: workbook is missing required tabs: " + ", ".join(missing_sheets),
          file=sys.stderr)
    sys.exit(2)

inp = wb["Inputs"]; rc = wb["Rent Capture"]; en = wb["Expense & NOI"]
eo = wb["Expense Overview"]; dq = wb["Delinquency"]; ho = wb["MTM"]
ma = wb["MTM Analysis"]; rp = wb["Renewal Pipeline"]; ug = wb["Unit Gap Analysis"]
ro = wb["Rate vs Occupancy"]; ld = wb["Lease Detail"]; fp = wb["Floorplan & Rollover"]
sc = wb["Scorecard"]

# ---- stale-workbook guard -------------------------------------------------
# If Excel has not recalculated, every formula reads None and the whole file
# would come out null. Fail here rather than publish that.
# probe the first month column (4 — column 3 is a GL reference in V37); a fixed
# far-right column would fall past the TTM column as months are added
missing, total = is_stale(None, [
    (rc, "Total rental income (accrual basis)", 4),
    (en, "Total operating expense", 4),
    (ho, "Holdover units", 3),
    (inp, "Total units", 3),
])
if missing == total:
    print(f"FATAL: no cached formula results ({missing}/{total} probes empty). The "
          "workbook was saved without recalculating — open it in Excel, let it "
          "recalculate, save, and re-run.", file=sys.stderr)
    sys.exit(2)
check(missing == 0, f"{missing} of {total} probe cells are empty — the workbook may "
                    "be partially stale", hard=False)

data = {}

# ---- meta / inputs -------------------------------------------------------
UNITS = scalar(inp, "Total units")
data["meta"] = {
    "property": scalar(inp, "Property"),
    "units": UNITS,
    "rentable_sqft": scalar(inp, "Total rentable sq ft"),
    "rent_roll_as_of": scalar(inp, "Rent roll as-of date"),
    "generated_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
    "source_workbook": os.path.basename(args.workbook),
    "source_reports": ["12-month accrual statement (T12)", "rent roll",
                       "delinquency summary", "renewal tracker",
                       "RealPage rate tracker (typed in)"],
    "note": ("Derived from the analyst workbook, which is fed by the same Yardi/RealPage "
             "reports the Drive pipeline collects. Refresh by pasting new reports into "
             "the workbook's grey Source tabs, recalculating in Excel, and re-running "
             "scripts/extract_landing.py."),
}
data["inputs"] = {
    # V37 moved the economic-occupancy target to Rate vs Occupancy and keeps a
    # physical target on Inputs; the scenario cards now run on an incremental
    # vacancy haircut rather than one-time make-ready/downtime costs.
    "target_econ_occupancy": scalar(ro, "Target economic occupancy"),
    "target_phys_occupancy": scalar(inp, "Target physical occupancy"),
    "incr_vacancy": scalar(inp, "Incremental vacancy applied to new run-rate"),
    "avg_phys_vacancy_ttm": scalar(inp, "Average physical vacancy, TTM (benchmark)"),
    "yardi_market_psf": scalar(inp, "Yardi market rent $/sqft (weighted)"),
    "inplace_psf": scalar(inp, "In-place rent $/sqft (weighted)"),
    "capture_rate": scalar(inp, "SELECTED capture rate - same unit, market at signing"),
    "turnover_cost_ttm": scalar(inp, "Turnover cost, TTM"),
    "mgmt_fee_pct": scalar(inp, "Management fee as % of revenue"),
    "make_ready_per_moveout": scalar(inp, "Direct make-ready cost per move-out"),
    "downtime_months": scalar(inp, "Downtime per move-out (months)"),
    "concession_per_lease": scalar(inp, "Concession per new lease ($)"),
    "leasing_cost_per_lease": scalar(inp, "Leasing and marketing cost per new lease ($)"),
    "cap_rate": scalar(rp, "Cap rate", col=3, value_col=4),
}

# ---- rent capture --------------------------------------------------------
rc_hdr = find_row(rc, "Line item")
rc_pct_hdr = find_row(rc, "AS % OF MARKET RENT POTENTIAL")
months, rc_cols, rc_ttm = month_axis(rc, rc_hdr)
S = lambda label, **kw: labelled_series(rc, label, rc_cols, **kw)
data["rent_capture"] = {
    "months": months,
    "market_potential": S("Market rent potential (Yardi)"),
    "loss_to_lease": S("Loss to lease (positive = in-place below market)"),
    "vacancy_loss": [-(v or 0) for v in S("Vacancy loss", before=rc_pct_hdr)],
    "employee_allowance": [-(v or 0) for v in S("Employee rent allowance", before=rc_pct_hdr)],
    "concessions": [-(v or 0) for v in S("Rental concessions", before=rc_pct_hdr)],
    "rental_income": S("Total rental income (accrual basis)"),
    "ltl_pct": S("Loss to lease", after=rc_pct_hdr),
    "capture_rate": S("Rent capture rate (accrued income / potential, before credit loss)"),
    "econ_occupancy": S("Economic occupancy"),
    "ttm": {
        "market_potential": scalar(rc, "Market rent potential (Yardi)", value_col=rc_ttm),
        "rental_income": scalar(rc, "Total rental income (accrual basis)", value_col=rc_ttm),
        "capture_rate": scalar(rc, "Rent capture rate (accrued income / potential, before credit loss)", value_col=rc_ttm),
        "econ_occupancy": scalar(rc, "Economic occupancy", value_col=rc_ttm),
        "ltl_pct": scalar(rc, "Loss to lease", value_col=rc_ttm, after=rc_pct_hdr),
    },
}

# ---- expense & NOI ------------------------------------------------------
en_hdr = find_row(en, "Line item")
en_norm_hdr = find_row(en, "NORMALISED VIEW - April 2026 real estate tax true-up smoothed",
                       exact=False, required=False) or en.max_row
en_months, en_cols, en_ttm = month_axis(en, en_hdr)
E = lambda label, **kw: labelled_series(en, label, en_cols, **kw)
data["expense_noi"] = {
    "months": en_months,
    "revenue": E("Total revenue"),
    "opex": E("Total operating expense", before=en_norm_hdr),
    "noi": E("Net operating income"),
    "opex_ratio": E("Operating expense ratio", before=en_norm_hdr),
    "noi_margin": E("NOI margin", before=en_norm_hdr),
    "controllable": E("Controllable"),
    "non_controllable": E("Non-controllable (taxes, insurance, mgmt fee)"),
    "ttm": {
        "revenue": scalar(en, "Total revenue", value_col=en_ttm),
        "opex": scalar(en, "Total operating expense", value_col=en_ttm, before=en_norm_hdr),
        "noi": scalar(en, "Net operating income", value_col=en_ttm),
        "operating_noi": scalar(en, "Operating NOI (excl. interest income)", value_col=en_ttm),
        "opex_ratio": scalar(en, "Operating expense ratio", value_col=en_ttm, before=en_norm_hdr),
        "noi_margin": scalar(en, "NOI margin", value_col=en_ttm, before=en_norm_hdr),
        "revenue_per_unit": scalar(en, "Revenue per unit", value_col=en_ttm),
        "opex_per_unit": scalar(en, "Operating expense per unit", value_col=en_ttm),
        "noi_per_unit": scalar(en, "NOI per unit", value_col=en_ttm),
    },
    "tax_note": ("April 2026 real estate tax true-up distorts that month; "
                 "TTM run-rate for taxes is ~172.5k/mo. Whether the April entry is a "
                 "one-off or an under-accrual catch-up is an open question — see the "
                 "tax-accrual readings."),
    # V37 adds an explicit open question: is the April posting a one-off
    # (Reading A) or a catch-up implying a higher run rate (Reading B)?
    "tax_accrual": {
        "monthly_accrual": scalar(en, "Monthly accrual in place (GL 510200-0001)"),
        "april_posting": scalar(en, "April 2026 posting"),
        "one_time_amount": scalar(en, "One-time amount in April 2026"),
        "reading_a_opex": scalar(en, "Reading A - April one-off: clean TTM opex"),
        "reading_a_ratio": scalar(en, "Reading A - clean TTM expense ratio"),
        "reading_a_noi": scalar(en, "Reading A - clean TTM NOI"),
        "reading_b_opex": scalar(en, "Reading B - Jul annualised opex"),
        "reading_b_ratio": scalar(en, "Reading B - Jul annualised expense ratio"),
        "reading_b_noi": scalar(en, "Reading B - Jul annualised NOI"),
    },
}

# ---- expense deep dive --------------------------------------------------
eo_hdr = find_row(eo, "Bucket")
data["expense_overview"] = {
    "period_note": cell(eo, find_row(eo, "Matching periods", exact=False), 2),
    "buckets": [dict(zip(["name", "p2025", "p2026", "change", "ttm", "per_unit",
                          "share", "status", "read"], r))
                for r in block(eo, eo_hdr, cols=range(2, 11))],
    "opportunities": [dict(zip(["name", "basis", "saving", "value", "difficulty"], r))
                      for r in block(eo, find_row(eo, "Opportunity"), cols=range(2, 7),
                                     stop_on_blank=False, max_rows=10,
                                     keep=lambda v: v[0] and v[2] is not None)],
    "revenue_compare": [dict(zip(["name", "saving", "value"], r))
                        for r in block(eo, find_row(eo, "For comparison - the revenue side",
                                                    exact=False),
                                       cols=[2, 4, 5], stop_on_blank=False, max_rows=8,
                                       # the last line carries a value but no annual
                                       # saving, so accept either
                                       keep=lambda v: v[0] and (v[1] is not None
                                                                or v[2] is not None))],
}

# ---- delinquency -------------------------------------------------------
def _dq_as_of():
    """Pull the snapshot date out of the tab's own intro sentence."""
    import re as _re
    txt = str(cell(dq, find_row(dq, "Resident balances", exact=False), 2) or "")
    m = _re.search(r"as of (\d{1,2}) (\w+) (\d{4})", txt)
    if not m:
        warn("delinquency as-of date not found in the tab's intro text")
        return None
    mon = ["january", "february", "march", "april", "may", "june", "july",
           "august", "september", "october", "november", "december"]
    return f"{m.group(3)}-{mon.index(m.group(2).lower()) + 1:02d}-{int(m.group(1)):02d}"

dq_aging_hdr = find_row(dq, "Bucket")
dq_detail_hdr = find_row(dq, "Unit")
# Resident names are deliberately NOT extracted. The published JSON sits on a
# public URL, and a surname next to a unit number and an amount past due is
# identifiable personal financial data. No chart needs it — the aging bars use
# amounts only — so the unit number alone identifies the row for anyone who
# needs to act on it. Column 3 (the name) is skipped in the column list below.
detail = block(dq, dq_detail_hdr, cols=[2, 4, 5, 6, 7, 8], stop_on_blank=False, max_rows=200,
               keep=lambda v: isinstance(v[1], (int, float)) and v[1] > 0
                              and not str(v[0]).upper().startswith("NONRES"))
detail.sort(key=lambda r: -r[1])
data["delinquency"] = {
    "as_of": _dq_as_of(),
    "units_with_balance": scalar(dq, "Units with a balance owed"),
    "gross_owed": scalar(dq, "Total owed (gross)"),
    "credits": scalar(dq, "Credits and prepayments"),
    "net": scalar(dq, "Net position"),
    "pct_month_rent": scalar(dq, "Total owed as % of one month rental income"),
    "occupied_units": scalar(dq, "Occupied units"),
    "share_with_balance": scalar(dq, "Share of residents carrying a balance"),
    "retail_balance": scalar(dq, "Retail balance outstanding"),
    "retail_over90": scalar(dq, "of which over 90 days", exact=False)
                     if find_row(dq, "of which over 90 days", exact=False, required=False) else None,
    "total_all": scalar(dq, "TOTAL DELINQUENCY - residential and retail"),
    "aging": [dict(zip(["bucket", "amount", "pct_owed", "pct_rent", "status"], r))
              for r in block(dq, dq_aging_hdr, cols=range(2, 7))],
    "top": [dict(zip(["unit", "owed", "d30", "d60", "d90", "over90"],
                     [str(r[0])] + r[1:])) for r in detail[:12]],
}

# V37 tracks unit 531 separately (it is also the rent-override unit on Inputs).
# The resident's name sits beside the "Resident" label and is deliberately not
# read — the unit number identifies the case. The balance labels carry their
# snapshot dates, so they are matched loosely and kept as labels.
u531_hdr = find_row(dq, "UNIT 531 - COLLECTIONS STATUS", exact=False, required=False)
if u531_hdr:
    b1 = find_row(dq, "balance owed,", after=u531_hdr, exact=False)
    b2 = find_row(dq, "balance owed,", after=b1 + 1, exact=False)
    data["delinquency"]["unit_531"] = {
        "status": scalar(dq, "Resident status per Yardi", value_col=4, after=u531_hdr),
        "memo": scalar(dq, "Delinquency memo from the property", value_col=4,
                       after=u531_hdr),
        "balance_prior": {"label": cell(dq, b1, 2), "amount": cell(dq, b1, 4)},
        "balance_latest": {"label": cell(dq, b2, 2), "amount": cell(dq, b2, 4)},
        "new_charges_week": scalar(dq, "of which new charges posted in the week",
                                   value_col=4, after=u531_hdr, exact=False),
        "paid_week": scalar(dq, "amount paid during the week", value_col=4,
                            after=u531_hdr, exact=False),
        "aged_over_30": scalar(dq, "Aged 31-60 / 61-90 / over 90 days",
                               value_col=4, after=u531_hdr),
    }
else:
    data["delinquency"]["unit_531"] = None

# ---- holdovers (the MTM tab) --------------------------------------------
# V37 renamed the tab and reworked the repricing model: each unit now carries
# its own applied increase and a vacate flag (Y = re-lease at market), and the
# roll-up charges a recurring incremental-vacancy haircut on the new run-rate
# instead of one-time make-ready/downtime costs.
ho_dist_hdr = find_row(ho, "Gap versus market")
ho_unit_hdr = find_row(ho, "Rank")
ho_units = block(ho, ho_unit_hdr, cols=[2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12,
                                        13, 14, 15, 16, 17, 18, 19, 20],
                 keep=lambda v: isinstance(v[1], (int, float, str)) and v[3] is not None)
data["holdovers"] = {
    "summary": {"units": scalar(ho, "Holdover units"),
                "inplace_mo": scalar(ho, "Their in-place rent, monthly"),
                "market_mo": scalar(ho, "Their market rent, monthly"),
                "gap_mo": scalar(ho, "Gap, monthly"),
                "gap_yr": scalar(ho, "Gap, annualised"),
                "cohort_below_mkt": scalar(ho, "Cohort is below market by"),
                "property_below_mkt": scalar(ho, "Whole property is below market by"),
                "worse_than_property": scalar(ho, "Holdovers are worse than the property",
                                              exact=False),
                "at_or_above_market": scalar(ho, "Units at or above market"),
                "share_of_ltl": scalar(ho, "Share of the total loss to lease they represent"),
                "share_of_units": scalar(ho, "Share of total units at the property")},
    "distribution": [dict(zip(["band", "units", "gap_yr", "share", "flag"], r))
                     for r in block(ho, ho_dist_hdr, cols=range(2, 7))],
    "threshold_default": scalar(ho, "Minimum gap % to include"),
    "units": [dict(zip(["rank", "unit", "type", "sqft", "inplace", "market", "gap_mo",
                        "gap_yr", "pct_below", "expired", "months_expired",
                        "inplace_psf", "market_psf", "spread_psf",
                        "inc_applied", "vacate", "new_rent", "incr_mo", "incr_yr"],
                       [r[0], str(r[1])] + r[2:15]
                       + [str(r[15]).upper() == "Y"] + r[16:])) for r in ho_units],
    "workbook_reprice": {
        "units_in_scope": scalar(ho, "Holdover units in scope"),
        "units_repriced": scalar(ho, "Units repriced in place"),
        "units_vacating": scalar(ho, "Units vacating"),
        "increase_wtd": scalar(ho, "Weighted average rent increase applied"),
        "share_of_gap_captured": scalar(ho, "Implied share of the gap captured"),
        "incremental_yr": scalar(ho, "Incremental rent, annualised"),
        "less_vacancy": scalar(ho, "Less: incremental vacancy", exact=False),
        "less_mgmt_fee": scalar(ho, "Less: management fee", exact=False),
        "recurring_noi": scalar(ho, "Recurring additional NOI"),
        "value": scalar(ho, "NET VALUE AT SALE"),
        "value_per_unit": scalar(ho, "Value per holdover unit"),
        "new_lease_discount": scalar(ho, "Discount to Market for new lease",
                                     col=15, value_col=16),
        "implied_vacancy": scalar(ho, "Implied Vacancy from this batch",
                                  col=15, value_col=16, exact=False)},
    # Three systems disagree about who is month-to-month; the reconciliation
    # lives on this tab and the counts are worth publishing.
    "three_systems": {
        "property_status_report": cell(ho, find_row(ho, "Property Status report",
                                                    exact=False,
                                                    after=find_row(ho, "Three systems",
                                                                   exact=False)), 4),
        "renewal_tracker_flagged": cell(ho, find_row(ho, "Renewal tracker, units flagged MTM"), 4),
        "this_model": cell(ho, find_row(ho, "This model, holdovers by rent roll test"), 4),
        "tracker_mtm_not_holdover": [
            str(v) for v in series(ho, find_row(ho, "Units the renewal tracker flags MTM",
                                                exact=False), range(4, 20))
            if v is not None],
    },
}

# ---- MTM analysis (tracker reconciliation) -------------------------------
# The tab reproduces the renewal tracker's MTM list, which carries resident
# names. Only the aggregate counts are read; the roster itself is not — names
# must not reach the published JSON (see check_no_pii.py).
ma_tot = find_row(ma, "Totals")
data["mtm_analysis"] = {
    "source": cell(ma, find_row(ma, "Source:", exact=False), 2),
    "tracker_total": cell(ma, ma_tot, 3),
    "status_counts": [dict(zip(["status", "units", "share"], r))
                      for r in block(ma, ma_tot, cols=range(2, 5),
                                     keep=lambda v: isinstance(v[1], (int, float)))],
    "reconciliation": {
        "tracker_mtm_units": scalar(ma, "Units on the tracker MTM tab", value_col=4),
        "still_mtm_on_rent_roll": scalar(ma, "still month-to-month on the rent roll",
                                         value_col=4, exact=False),
        "not_mtm_lease_running": scalar(ma, "NOT month-to-month", value_col=4,
                                        after=find_row(ma, "WHAT THE ADDED COLUMN SHOWS"),
                                        exact=False),
        "not_found_on_rent_roll": scalar(ma, "not found on the rent roll", value_col=4,
                                         exact=False),
        "share_tracker_wrong": scalar(ma, "Share of the tracker list that is wrong",
                                      value_col=4, exact=False),
        "holdovers_on_rent_roll": scalar(ma, "Holdovers on the", value_col=4, exact=False),
        "holdovers_on_tracker": scalar(ma, "appearing on the tracker MTM tab",
                                       value_col=4, exact=False),
        "holdovers_missing_from_tracker": scalar(ma, "missing from the tracker MTM tab",
                                                 value_col=4, exact=False),
        "share_holdovers_unidentified": scalar(ma, "Share of the holdover cohort not",
                                               value_col=4, exact=False),
    },
}

# ---- renewal pipeline --------------------------------------------------
# V37 moved this tab's labels to column C, extended the table to Q2 2027 with
# per-month NOI math, and replaced the one-time lease-up costs with a recurring
# incremental-vacancy haircut on the new run-rate (see the SCENARIO note on the
# tab). It also added the offers actually issued and a model-vs-tracker
# calibration.
RPC = dict(col=3, value_col=4)          # labels in C, values in D
rp_hdr = find_row(rp, "Month", col=3)
rp_rows = block(rp, rp_hdr, cols=list(range(3, 25)) + [26, 27, 28], label_col=3,
                keep=lambda v: parse_month(v[0]) is not None)
data["renewal"] = {
    "months": [{"month": parse_month(r[0]), "inc": r[1], "ret": r[2], "expiring": r[3],
                "on_notice": r[4], "var_units": r[5], "sqft": r[6], "inplace_var": r[7],
                "market_var": r[8], "inplace_on": r[9], "market_on": r[10],
                "capped_var": r[11], "capped_on": r[12], "renewing": r[13],
                "moveouts": r[14], "new_rent": r[15], "current_rent": r[16],
                "incr_mo": r[17], "incr_yr": r[18], "less_vacancy": r[19],
                "recurring_noi": r[20], "value_at_cap": r[21],
                "key_from": r[22], "key_to": r[23], "basis": r[24]} for r in rp_rows],
    "assumptions": {
        "on_notice_increase": scalar(rp, "On-notice units - renewal increase", **RPC),
        "on_notice_retention": scalar(rp, "On-notice units - retention rate", **RPC),
        "make_ready": scalar(rp, "Direct make-ready cost per move-out", exact=False, **RPC),
        "market_psf_assumption": scalar(rp, "Market rent assumption", exact=False, **RPC),
        "market_scaling": scalar(rp, "Scaling factor applied to market", exact=False, **RPC),
    },
    # the TTM operating NOI moved off this tab; the uplift denominator now
    # comes from the Expense & NOI TTM block extracted above
    "ttm_operating_noi": data["expense_noi"]["ttm"]["operating_noi"],
    "ftm_value": scalar(rp, "FTM Value", exact=False, required=False, **RPC),
    "workbook_rollup": {
        "incremental_yr": scalar(rp, "Incremental rent, annualised", **RPC),
        "less_vacancy": scalar(rp, "Less: incremental vacancy", exact=False, **RPC),
        "less_mgmt_fee": scalar(rp, "Less: management fee on collected revenue", **RPC),
        "recurring_noi": scalar(rp, "Recurring change in NOI", **RPC),
        "value_at_cap": scalar(rp, "Value of the recurring NOI change at the cap rate",
                               **RPC),
        "total_value_at_sale": scalar(rp, "Total Value at Sale", exact=False,
                                      required=False, **RPC)},
    "scenario_note": cell(rp, find_row(rp, "SCENARIO - stabilised sensitivity",
                                       col=3, exact=False) + 1, 3),
    "model_note": ("Recomputed live from unit level. Stabilised sensitivity, not a "
                   "forecast: every move-out re-leases at the market assumption, and "
                   "vacancy/downtime are charged as a recurring incremental-vacancy "
                   "haircut on the new run-rate rather than one-time costs."),
}

# what is actually being offered, straight from the tracker's monthly sheets
rp_offers_hdr = find_row(rp, "Month", col=3,
                         after=find_row(rp, "ACTUAL RENEWAL OFFERS", col=3, exact=False))
data["renewal"]["offers"] = {
    "months": [dict(zip(["month", "leases", "current_rent", "offered_rent", "wtd_increase",
                         "ltl_before", "ltl_after", "gap_surviving"],
                        [parse_month(r[0])] + r[1:]))
               for r in block(rp, rp_offers_hdr, cols=range(3, 11), label_col=3,
                              keep=lambda v: parse_month(v[0]) is not None)],
    "total": dict(zip(["leases", "current_rent", "offered_rent", "wtd_increase"],
                      series(rp, find_row(rp, "Total / weighted", col=3), range(4, 8)))),
    "calibration": {
        "model_inplace": scalar(rp, "Model: in-place rent on the variab", exact=False, **RPC),
        "model_renewal_rent": scalar(rp, "Model: renewal rent after", exact=False, **RPC),
        "model_rate": scalar(rp, "Model: renewal rent after", exact=False,
                             col=3, value_col=5),
        "tracker_current": scalar(rp, "Tracker: current rent on the same", exact=False, **RPC),
        "tracker_offered": scalar(rp, "Tracker: rent actually offered", **RPC),
        "tracker_rate": scalar(rp, "Tracker: rent actually offered", col=3, value_col=5),
        "model_above_actual_pts": scalar(rp, "Model above actual, in points",
                                         col=3, value_col=5),
        "ratio_actual_to_model": scalar(rp, "Calibration ratio (actual / model)",
                                        col=3, value_col=5),
        "implied_rate_on_15pct_policy": scalar(rp, "Implied effective rate on a 15%",
                                               exact=False, col=3, value_col=5),
    },
}

# ---- unit-level gap data ----------------------------------------------
ug_hdr = find_row(ug, "Unit")
ug_rows = block(ug, ug_hdr, cols=[2, 3, 4, 5, 6, 7, 8, 12, 13, 21],
                stop_on_blank=False, max_rows=600,
                # a real unit has a numeric size and market rent; the summary
                # blocks further down the sheet put labels in the same column
                keep=lambda v: isinstance(v[2], (int, float)) and v[2]
                               and isinstance(v[3], (int, float)) and v[3])
data["units"] = [{"unit": str(r[0]), "type": r[1], "sqft": r[2], "market": r[3],
                  "inplace": r[4], "gap": r[5], "gap_pct": r[6], "expiry_key": r[7],
                  "status": r[8], "market_assum": r[9]} for r in ug_rows]

# ---- rate vs occupancy ------------------------------------------------
ro_grid_hdr = find_row(ro, "Loss to lease", after=find_row(ro, "SCENARIO GRID", exact=False))
grid_vac_cols = [c for c in range(3, ro.max_column + 1)
                 if isinstance(cell(ro, ro_grid_hdr, c), (int, float))]
grid_rows = block(ro, ro_grid_hdr, cols=[2] + grid_vac_cols,
                  keep=lambda v: isinstance(v[0], (int, float)))
data["rate_occ"] = {
    "market_rent_yr": scalar(ro, "Market rent, annualised (rent roll 14 Jul 2026)", exact=False),
    "cur_ltl": scalar(ro, "Current loss to lease"),
    "cur_vacancy": scalar(ro, "Current vacancy (market rent on vacant units)"),
    "inplace_yr": scalar(ro, "Current in-place rent, annualised"),
    "target_occupancy": scalar(ro, "Target economic occupancy"),
    "breakeven_ltl": scalar(ro, "Loss to lease required to break even"),
    "scenario_ltl": scalar(ro, "Scenario: loss to lease narrows to"),
    "scenario_vac": scalar(ro, "Scenario: economic vacancy rises to"),
    "scenario_recurring_noi": scalar(ro, "Recurring change in NOI"),
    "scenario_net_value": scalar(ro, "Net value change"),
    "ttm_value_today": scalar(ro, "TTM Value Today"),
    "grid_vac": [cell(ro, ro_grid_hdr, c) for c in grid_vac_cols],
    "grid_ltl": [r[0] for r in grid_rows],
    "grid": [r[1:] for r in grid_rows],
}

# ---- leasing / trade-outs ---------------------------------------------
ld_hdr = find_row(ld, "Lease date")
ld_rows = block(ld, ld_hdr, cols=[2, 3, 4, 5, 6, 7, 8, 9, 11, 16, 17],
                keep=lambda v: v[0] is not None and str(v[0])[:4].isdigit())
band_hdr = find_row(ld, "Band")
data["leasing"] = {
    "leases": [{"date": r[0], "unit": str(r[1]), "sqft": r[2], "term": r[3], "rent": r[4],
                "psf": r[5], "prior": r[6], "tradeout": r[7], "term_type": r[8],
                "capture": r[9], "status": r[10]} for r in ld_rows],
    "summary": {
        "executed": scalar(ld, "Leases (executed)"),
        "wtd_psf": scalar(ld, "Weighted $/sqft - all leases"),
        "wtd_tradeout": scalar(ld, "Dollar-weighted trade-out"),
        # this label appears twice (psf block and trade-out block) — scope it
        "avg_tradeout_exec": scalar(ld, "executed leases only",
                                    after=find_row(ld, "Dollar-weighted trade-out")),
        "below_prior": scalar(ld, "Leases below prior rent"),
        "capture_at_signing": scalar(ld, "CAPTURE RATE - same unit, market at signing", exact=False),
        "capture_vs_july": scalar(ld, "same unit vs July market only", exact=False)},
    "bands": [dict(zip(["band", "range", "leases", "sqft", "achieved_psf"], r))
              for r in block(ld, band_hdr, cols=range(2, 7),
                             keep=lambda v: isinstance(v[0], (int, float)))],
    "renewal_activity": {
        "tracker_date": scalar(ld, "Tracker snapshot date"),
        "renewals_signed": scalar(ld, "Renewals signed"),
        "new_signed": scalar(ld, "New leases signed"),
        "renewal_rate_psf": scalar(ld, "Renewal rate, blended ($/sqft)"),
        "avg_increase": scalar(ld, "Average renewal increase"),
        "leased_pct": scalar(ld, "Leased %"),
        "spread_psf": scalar(ld, "Spread - new lease less renewal ($/sqft)"),
        "renewals_below_new": scalar(ld, "Renewals priced below new leases")},
}

# ---- floorplans + rollover -------------------------------------------
fp_hdr = find_row(fp, "Floorplan")
fp_rows = block(fp, fp_hdr, cols=range(2, 12), keep=lambda v: isinstance(v[1], (int, float)))
fp_total_row = find_row(fp, "Total", after=fp_hdr)
roll_hdr = find_row(fp, "Month", after=find_row(fp, "ROLLOVER SCHEDULE", exact=False))
data["floorplans"] = {
    "rows": [dict(zip(["plan", "units", "sqft", "avg_sqft", "market", "inplace",
                       "gap_mo", "gap_pct", "market_psf", "inplace_psf"], r))
             for r in fp_rows],
    "total": dict(zip(["units", "sqft", "avg_sqft", "market", "inplace", "gap_mo",
                       "gap_pct", "market_psf", "inplace_psf"],
                      [cell(fp, fp_total_row, c) for c in range(3, 12)])),
}
data["rollover"] = [dict(zip(["month", "units", "pct", "sqft", "inplace", "market",
                              "uncaptured", "cum_units", "cum_pct"], r))
                    for r in block(fp, roll_hdr, cols=range(2, 11),
                                   keep=lambda v: isinstance(v[1], (int, float)))]

# ---- scored insights (the workbook's own Scorecard tab) -------------------
# New in V37: a scored read of the data plus the open questions. This is the
# property-level insights table, not the portfolio KPI scorecard (which stays
# in docs/scorecard.json from its own workbook).
sc_hdr = find_row(sc, "Metric")
sc_flags_hdr = find_row(sc, "Item", after=find_row(sc, "RED FLAGS AND OPEN QUESTIONS"))
data["insights"] = {
    "metrics": [dict(zip(["name", "value", "status", "meaning"], r))
                for r in block(sc, sc_hdr, cols=range(2, 6))],
    "flags": [dict(zip(["item", "status", "why", "settle"], r))
              for r in block(sc, sc_flags_hdr, cols=[2, 3, 5, 6])],
}

# =========================== VALIDATION ===============================
# Each check guards a way the extraction could go quietly wrong.
V = []
def record(name, ok, detail, hard=True):
    V.append((name, ok, detail))
    check(ok, f"{name}: {detail}", hard=hard)

record("month axis aligned",
       data["rent_capture"]["months"] == data["expense_noi"]["months"],
       f"Rent Capture has {len(months)} months ending {months[-1]}, "
       f"Expense & NOI has {len(en_months)} ending {en_months[-1]}")

record("month count sane", len(months) >= 12,
       f"{len(months)} month columns found (expected 12 or more)")

n_units = len(data["units"])
record("unit count matches Inputs", n_units == UNITS,
       f"{n_units} unit rows read, Inputs says {UNITS} total units")

record("no summary rows in units",
       all(u["sqft"] and u["market"] for u in data["units"]),
       "every unit row has a numeric sq ft and market rent")

tie_rc = labelled_series(rc, "Check vs statement (should be nil)", rc_cols)
tie_rev = labelled_series(en, "Check vs statement (should be nil)", en_cols,
                          before=find_row(en, "OPERATING EXPENSES"))
tie_opex = labelled_series(en, "Check vs statement (should be nil)", en_cols,
                           after=find_row(en, "OPERATING EXPENSES"))
worst_tie = max(abs(v or 0) for v in tie_rc + tie_rev + tie_opex)
record("statement tie-outs nil", worst_tie < 1,
       f"largest tie-out variance is {worst_tie:.2f} (must be under 1.00)")

hd = data["holdovers"]
record("holdover rows match summary", len(hd["units"]) == hd["summary"]["units"],
       f"{len(hd['units'])} holdover rows read, summary says {hd['summary']['units']}")

record("no Total row in holdovers",
       all(norm(u["unit"]) != "total" for u in hd["units"]),
       "no 'Total' row leaked into the holdover unit list")

dist_total = sum(d["gap_yr"] or 0 for d in hd["distribution"])
record("holdover distribution ties to gap",
       abs(dist_total - (hd["summary"]["gap_yr"] or 0)) < 1,
       f"distribution sums to {dist_total:,.0f} vs summary gap {hd['summary']['gap_yr']:,.0f}")

fpt = data["floorplans"]["total"]
record("floorplan total matches Inputs", fpt["units"] == UNITS,
       f"floorplan total is {fpt['units']} units, Inputs says {UNITS}")
record("floorplan rows sum to total",
       sum(r["units"] or 0 for r in data["floorplans"]["rows"]) == fpt["units"],
       f"{len(data['floorplans']['rows'])} floorplan rows sum to "
       f"{sum(r['units'] or 0 for r in data['floorplans']['rows'])} vs total {fpt['units']}")

ag = sum(a["amount"] or 0 for a in data["delinquency"]["aging"])
record("delinquency aging ties to gross",
       abs(ag - (data["delinquency"]["gross_owed"] or 0)) < 1,
       f"aging buckets sum to {ag:,.2f} vs gross owed {data['delinquency']['gross_owed']:,.2f}")

ex = [l for l in data["leasing"]["leases"] if l["status"] == "Executed"]
record("executed lease count matches summary",
       len(ex) == data["leasing"]["summary"]["executed"],
       f"{len(ex)} executed leases read, summary says {data['leasing']['summary']['executed']}")

# The renewal card recomputes from unit level; if no units match a month's
# expiry-key window the card would silently show zero incremental rent.
cur_units = [u for u in data["units"] if u["status"] == "Current"]
def in_window(m):
    return [u for u in cur_units if m["key_from"] <= u["expiry_key"] < m["key_to"]]
empty_months = [m["month"] for m in data["renewal"]["months"] if not in_window(m)]
record("renewal months resolve to units", not empty_months,
       "no units matched expiry months " + ", ".join(empty_months) if empty_months
       else f"all {len(data['renewal']['months'])} months matched unit rows")

# Reproduce the workbook's own roll-up from the unit-level data we emit, using
# the V37 model: variable pool renews at min(in-place x (1+inc), market at
# assumption), move-outs and on-notice units re-lease at the market assumption,
# and a recurring incremental-vacancy haircut applies to the new run-rate.
mgmt = data["inputs"]["mgmt_fee_pct"]
vac = data["inputs"]["incr_vacancy"] or 0
incr = new_run = worst_cap = 0.0
for m in data["renewal"]["months"]:
    capped = sum(min(u["inplace"] * (1 + m["inc"]), u["market_assum"])
                 for u in in_window(m))
    worst_cap = max(worst_cap, abs(capped - (m["capped_var"] or 0)))
    new = capped * m["ret"] + m["market_var"] * (1 - m["ret"]) + m["market_on"]
    incr += (new - (m["inplace_var"] + m["inplace_on"])) * 12
    new_run += new * 12
wb_ru = data["renewal"]["workbook_rollup"]
wb_incr = wb_ru["incremental_yr"] or 0
record("renewal model reproduces workbook",
       abs(incr - wb_incr) < max(1000, abs(wb_incr) * 0.002),
       f"recomputed incremental rent {incr:,.0f} vs workbook {wb_incr:,.0f} "
       f"(worst per-month cap variance {worst_cap:,.0f})")

month_incr = sum(m["incr_yr"] or 0 for m in data["renewal"]["months"])
record("renewal months sum to roll-up", abs(month_incr - wb_incr) < 1,
       f"per-month incremental sums to {month_incr:,.0f} vs roll-up {wb_incr:,.0f}")

rec_noi = (incr - vac * new_run) * (1 - mgmt)
wb_rec = wb_ru["recurring_noi"] or 0
record("renewal vacancy haircut reproduces NOI",
       abs(rec_noi - wb_rec) < max(1000, abs(wb_rec) * 0.002),
       f"recomputed recurring NOI {rec_noi:,.0f} vs workbook {wb_rec:,.0f} "
       f"({vac:.1%} vacancy on the new run-rate, then the management fee)")

offers = data["renewal"]["offers"]
record("renewal offers table read", len(offers["months"]) >= 4
       and offers["total"]["leases"] == sum(o["leases"] for o in offers["months"]),
       f"{len(offers['months'])} months, {offers['total']['leases']} offers "
       "(total row ties to the month rows)")

hr = data["holdovers"]["workbook_reprice"]
unit_incr = sum(u["incr_yr"] or 0 for u in data["holdovers"]["units"])
record("holdover unit rows sum to reprice roll-up",
       abs(unit_incr - (hr["incremental_yr"] or 0)) < 1,
       f"unit-level incremental sums to {unit_incr:,.0f} vs "
       f"roll-up {hr['incremental_yr'] or 0:,.0f}")
n_vac = sum(1 for u in data["holdovers"]["units"] if u["vacate"])
record("holdover vacate flags match roll-up", n_vac == hr["units_vacating"],
       f"{n_vac} units flagged Y vs roll-up {hr['units_vacating']}")

record("insights scorecard read",
       len(data["insights"]["metrics"]) >= 10 and len(data["insights"]["flags"]) >= 3
       and all(m["status"] in ("GOOD", "WATCH", "FLAG")
               for m in data["insights"]["metrics"]),
       f"{len(data['insights']['metrics'])} scored metrics, "
       f"{len(data['insights']['flags'])} open questions")

nulls = [k for k, v in data["inputs"].items() if v is None]
record("all inputs resolved", not nulls,
       "inputs with no value: " + ", ".join(nulls) if nulls else "all inputs found",
       hard=False)

# ---- report --------------------------------------------------------------
print(f"{'check':38} result")
print("-" * 78)
for name, ok, detail in V:
    print(f"{name:38} {'PASS' if ok else 'FAIL'}  {detail}")

if FATAL:
    print("\nFATAL — not writing " + args.out, file=sys.stderr)
    for m in FATAL:
        print("  * " + m, file=sys.stderr)
    print("\nThe workbook layout moved in a way that would put wrong numbers on the "
          "dashboard. Fix the workbook or update the anchors in this script.",
          file=sys.stderr)
    sys.exit(1)

with open(args.out, "w") as f:
    json.dump(data, f, separators=(",", ":"))

print(f"\nwrote {args.out} ({os.path.getsize(args.out)} bytes)")
print(f"  {len(data['units'])} units · {len(months)} months ({months[0]}..{months[-1]}) "
      f"· {len(data['holdovers']['units'])} holdovers · {len(ex)} executed leases "
      f"· {len(data['rollover'])} rollover rows")
if WARN:
    print("\nwarnings:")
    for m in WARN:
        print("  ! " + m)
    if not args.allow_warnings:
        sys.exit(3)
