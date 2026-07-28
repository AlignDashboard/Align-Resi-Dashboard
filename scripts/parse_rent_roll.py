"""
parse_rent_roll.py
------------------
Parses a Yardi rent roll export (the Rent Roll tab of the SPV PM Deliverable
Package) into per-unit records plus the totals the dashboard needs.

Anchors on the header labels rather than row/column positions, because the
export moves: units are added, columns are inserted, and the header itself is
split across two rows ("Market" above "Rent"). The report's own Total and
Summary Groups rows are used as a tie-out — if the parsed units do not sum to
what the report says, that is raised rather than returned.

Two traps this handles, both of which silently corrupt a naive read:

  * The unit rows are followed by a "Future Residents/Applicants" section of
    applicants who are not part of the unit count. Reading past that marker
    inflates the property.
  * "Total"/"Summary Groups" rows sit immediately below the data with values in
    the same columns, so they look like units.

Usage:
    from parse_rent_roll import parse
    result = parse("path/to/rent_roll.xlsx")
"""
import json
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from xlsx_anchors import (LayoutError, header_map, norm,  # noqa: E402
                          rows_until)

# field -> pattern matched against the joined header text
COLUMNS = {
    "unit": r"^unit$",
    "unit_type": r"^unit type$",
    "sqft": r"sq ?ft",
    "resident_code": r"^resident$",
    "resident_name": r"^name$",
    "market_rent": r"^market rent$",
    "actual_rent": r"^actual rent$",
    "resident_deposit": r"^resident deposit$",
    "other_deposit": r"^other deposit$",
    "move_in": r"^move in$",
    "lease_expiration": r"^lease expiration$",
    "move_out": r"^move out$",
    "balance": r"^balance$",
}

# Rows that end the current-resident section. Everything after is a different
# population (applicants) or a summary.
STOP = [r"^future residents", r"^total\b", r"^summary groups", r"^totals?:",
        r"^occupied units", r"^total vacant", r"^total non rev"]


def _text_cells(ws, max_row=30, max_col=None):
    out = []
    for r in range(1, min(ws.max_row, max_row) + 1):
        for c in range(1, min(ws.max_column, max_col or ws.max_column) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                out.append(v.strip())
    return out


def _property(ws):
    """('The Landing', 'p0005611') from a line like 'The Landing (p0005611)'."""
    for t in _text_cells(ws):
        m = re.match(r"^(.{2,60}?)\s*\(([A-Za-z0-9._-]{4,20})\)\s*$", t)
        if m and not m.group(1).lower().startswith("column "):
            return m.group(1).strip(), m.group(2).strip()
    return None, None


def _as_of(ws):
    for t in _text_cells(ws):
        m = re.search(r"as of\s*=?\s*(\d{1,2})/(\d{1,2})/(\d{4})", t, re.I)
        if m:
            return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
        m = re.search(r"as of\s*=?\s*(\d{4})-(\d{2})-(\d{2})", t, re.I)
        if m:
            return m.group(0).split("=")[-1].strip()[-10:]
    return None


def _report_totals(ws, fields):
    """The report's own Total row, used purely as a tie-out."""
    unit_col = fields["unit"]
    for r in range(1, ws.max_row + 1):
        row_text = " ".join(norm(ws.cell(row=r, column=c).value)
                            for c in range(1, min(ws.max_column, unit_col + 4) + 1))
        if re.search(r"\btotal\b", row_text) and not re.search(r"total (vacant|non rev)", row_text):
            mk = ws.cell(row=r, column=fields["market_rent"]).value
            ac = ws.cell(row=r, column=fields["actual_rent"]).value
            if isinstance(mk, (int, float)) and isinstance(ac, (int, float)):
                return {"row": r, "market_rent": float(mk), "actual_rent": float(ac)}
    return None


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").replace("$", "").strip()
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _pick_sheet(wb):
    """The rent roll tab, whatever the workbook calls it."""
    for name in wb.sheetnames:
        if "rent roll" in name.lower():
            return wb[name]
    return wb[wb.sheetnames[0]]


def parse(path, strict=True):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _pick_sheet(wb)
    fields, header_row = header_map(ws, COLUMNS)

    units, stopped = rows_until(
        ws, header_row + 1, fields, STOP, "unit",
        # a real unit row has a unit id and a numeric size and market rent
        keep=lambda rec: rec["unit"] not in (None, "")
                         and _num(rec["sqft"]) is not None
                         and _num(rec["market_rent"]) is not None)

    for u in units:
        u["unit"] = str(u["unit"]).strip()
        for k in ("sqft", "market_rent", "actual_rent", "resident_deposit",
                  "other_deposit", "balance"):
            u[k] = _num(u.get(k))
        u["occupied"] = bool(u.get("resident_code")) and (u.get("actual_rent") or 0) > 0
        u["on_notice"] = bool(u.get("move_out"))

    name, code = _property(ws)
    totals = {
        "units": len(units),
        "sqft": round(sum(u["sqft"] or 0 for u in units), 2),
        "market_rent": round(sum(u["market_rent"] or 0 for u in units), 2),
        "actual_rent": round(sum(u["actual_rent"] or 0 for u in units), 2),
        "balance": round(sum(u["balance"] or 0 for u in units), 2),
    }

    checks, problems = [], []
    rep = _report_totals(ws, fields)
    if rep:
        for key in ("market_rent", "actual_rent"):
            delta = abs(totals[key] - rep[key])
            ok = delta < 0.02
            checks.append({"check": f"{key} ties to the report Total row", "ok": ok,
                           "parsed": totals[key], "report": rep[key],
                           "delta": round(delta, 2)})
            if not ok:
                problems.append(
                    f"{key}: units sum to {totals[key]:,.2f} but the report's Total row "
                    f"(row {rep['row']}) says {rep[key]:,.2f}")
    else:
        checks.append({"check": "report Total row found", "ok": False,
                       "note": "no Total row located — cannot tie out"})
        problems.append("no Total row found in the export to tie out against")

    as_of = _as_of(ws)
    checks.append({"check": "as-of date found", "ok": as_of is not None,
                   "note": as_of or "no as-of date in the export — the caller should "
                                   "fall back to the file's date"})
    checks.append({"check": "stopped at a section marker", "ok": stopped is not None,
                   "note": f"stopped at row {stopped[0]} ({stopped[1]!r})" if stopped
                           else "ran to the end of the sheet without a marker"})
    if stopped is None:
        problems.append("no section marker after the unit rows — applicants or "
                        "summary rows may have been read as units")

    if not units:
        problems.append("no unit rows found")

    if problems and strict:
        raise ValueError("rent roll did not tie out:\n  - " + "\n  - ".join(problems))

    return {
        "report_type": "rent_roll",
        "property": name,
        "property_code": code,
        "as_of": as_of,
        "source_file": os.path.basename(path),
        "sheet": ws.title,
        "header_row": header_row,
        "columns": {k: v for k, v in sorted(fields.items(), key=lambda kv: kv[1])},
        "units": units,
        "totals": totals,
        "checks": checks,
        "problems": problems,
    }


if __name__ == "__main__":
    out = parse(sys.argv[1])
    slim = dict(out)
    slim["units"] = out["units"][:3] + [f"... {len(out['units']) - 3} more"]
    print(json.dumps(slim, indent=2, default=str))
