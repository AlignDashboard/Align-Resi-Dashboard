#!/usr/bin/env python3
"""Extract the workbook's grey Source tabs into standalone report fixtures.

The grey tabs are pastes of the raw Yardi exports, so they carry the same
headers, section markers and total rows. Lifting one into its own file gives a
realistic stand-in for a Drive drop to test the parsers against, which is the
closest thing available until a real export is on hand.

Two honest caveats: the paste starts at column C rather than A on the rent roll,
and the analyst added annotation text to the right of the data. Both are exactly
the kind of drift the parsers are supposed to absorb, so testing against them is
useful — but it is not a substitute for one real export.

Usage: python scripts/make_report_fixtures.py <workbook.xlsx> [outdir]
"""
import os
import sys

import openpyxl

SRC = sys.argv[1]
OUT = sys.argv[2] if len(sys.argv) > 2 else "tests/fixtures"

# grey tab -> (fixture filename, sheet name in the fixture, columns to keep)
TABS = {
    "Source Rent Roll Jul": ("rent_roll_jul.xlsx", "Rent Roll", range(1, 16)),
    "Source Rent Roll Jun": ("rent_roll_jun.xlsx", "Rent Roll", range(1, 16)),
    "Source Delinquency": ("delinquency.xlsx", "Delinquency", range(1, 16)),
}


def main():
    os.makedirs(OUT, exist_ok=True)
    src = openpyxl.load_workbook(SRC, data_only=True)
    made = []
    for tab, (fname, sheet, cols) in TABS.items():
        if tab not in src.sheetnames:
            print(f"[skip] {tab!r} not in workbook")
            continue
        ws_in = src[tab]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet
        for r in range(1, ws_in.max_row + 1):
            for c in cols:
                v = ws_in.cell(row=r, column=c).value
                if v is not None:
                    ws.cell(row=r, column=c, value=v)
        path = os.path.join(OUT, fname)
        wb.save(path)
        made.append((tab, path, ws_in.max_row))
        print(f"[ok] {tab:24} -> {path}  ({ws_in.max_row} rows)")
    if not made:
        sys.exit("no fixtures written")


if __name__ == "__main__":
    main()
