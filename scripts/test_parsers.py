#!/usr/bin/env python3
"""Tests for the rent roll and delinquency parsers.

Expected values are taken from the analyst workbook's own computed tabs, so a
pass means the parser reproduces numbers a human has already reviewed — not just
that it runs. Negative cases mutate the fixture in the ways a real export drifts
and require the parser to raise rather than return a plausible wrong number.

Fixtures come from scripts/make_report_fixtures.py, which lifts the workbook's
grey Source tabs into standalone files. They are realistic stand-ins for a Drive
drop but not real exports — see that script's caveats.

Usage:
  python scripts/make_report_fixtures.py <workbook.xlsx> tests/fixtures
  python scripts/test_parsers.py tests/fixtures
"""
import os
import shutil
import sys
import tempfile

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import parse_delinquency  # noqa: E402
import parse_rent_roll  # noqa: E402
from xlsx_anchors import LayoutError, header_map  # noqa: E402

FIX = sys.argv[1] if len(sys.argv) > 1 else "tests/fixtures"
PASS, FAIL = [], []


def ok(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    print(f"  {'PASS' if cond else 'FAIL'}  {name}" + (f"  — {detail}" if detail else ""))


def raises(name, fn, expect=(ValueError, LayoutError), needle=None):
    try:
        fn()
        ok(name, False, "no exception raised")
    except expect as e:
        ok(name, needle is None or needle.lower() in str(e).lower(),
           str(e).replace("\n", " ")[:88])
    except Exception as e:  # noqa: BLE001
        ok(name, False, f"wrong exception type {type(e).__name__}: {e}")


def copy(path):
    tmp = tempfile.mkdtemp()
    dst = os.path.join(tmp, os.path.basename(path))
    shutil.copy(path, dst)
    return dst


# ---------------------------------------------------------------- rent roll
RR = os.path.join(FIX, "rent_roll_jul.xlsx")
if not os.path.exists(RR):
    sys.exit(f"fixture missing: {RR}\nRun scripts/make_report_fixtures.py first.")

print("\nrent roll parser")
rr = parse_rent_roll.parse(RR)
ok("property identified", (rr["property"], rr["property_code"]) == ("The Landing", "p0005611"),
   f'{rr["property"]} / {rr["property_code"]}')
ok("as-of date read", rr["as_of"] == "2026-07-14", str(rr["as_of"]))
ok("unit count matches the workbook", rr["totals"]["units"] == 263,
   f'{rr["totals"]["units"]} units (Inputs says 263)')
ok("square footage matches", rr["totals"]["sqft"] == 215452.0, f'{rr["totals"]["sqft"]:,}')
ok("market rent matches", rr["totals"]["market_rent"] == 1701103.0,
   f'{rr["totals"]["market_rent"]:,}')
ok("in-place rent matches the report", rr["totals"]["actual_rent"] == 1234474.0,
   f'{rr["totals"]["actual_rent"]:,} (workbook shows 1,240,079 after the unit-531 override)')
ok("resident balances match Unit Gap Analysis", rr["totals"]["balance"] == 89952.89,
   f'{rr["totals"]["balance"]:,}')
ok("all tie-outs pass", all(c["ok"] for c in rr["checks"]),
   ", ".join(c["check"] for c in rr["checks"] if not c["ok"]) or "all")

# The trap: applicants sit directly below the units under their own marker.
ok("applicants excluded", all(u["unit"] != "Future Residents/Applicants" for u in rr["units"]))
stop = next(c for c in rr["checks"] if c["check"] == "stopped at a section marker")
ok("stopped at the applicants marker", "future residents" in stop["note"], stop["note"])
ok("no summary rows read as units",
   all(isinstance(u["sqft"], float) and u["sqft"] > 0 for u in rr["units"]))
ok("unit ids are strings", all(isinstance(u["unit"], str) for u in rr["units"]))
ok("on-notice flag set from Move Out",
   sum(1 for u in rr["units"] if u["on_notice"]) > 0,
   f'{sum(1 for u in rr["units"] if u["on_notice"])} units on notice')

# negative: break the report's own total so the sum no longer agrees
bad = copy(RR)
wb = openpyxl.load_workbook(bad)
ws = wb["Rent Roll"]
fields, _ = header_map(ws, parse_rent_roll.COLUMNS)
tot = parse_rent_roll._report_totals(ws, fields)
ws.cell(row=tot["row"], column=fields["market_rent"], value=999999)
wb.save(bad)
raises("broken Total row is refused", lambda: parse_rent_roll.parse(bad), needle="tie out")

# negative: a unit row deleted -> no longer ties to the report total.
# hdr + 1 is the "Current/Notice/Vacant Residents" section marker, so the first
# real unit is below it; find it rather than assuming an offset.
bad = copy(RR)
wb = openpyxl.load_workbook(bad)
ws = wb["Rent Roll"]
fields, hdr = header_map(ws, parse_rent_roll.COLUMNS)
first_unit = next(r for r in range(hdr + 1, ws.max_row + 1)
                  if isinstance(ws.cell(row=r, column=fields["sqft"]).value, (int, float)))
ws.delete_rows(first_unit)
wb.save(bad)
raises("dropped unit row is refused", lambda: parse_rent_roll.parse(bad), needle="tie out")

# negative: header gone -> cannot guess the columns
bad = copy(RR)
wb = openpyxl.load_workbook(bad)
ws = wb["Rent Roll"]
fields, hdr = header_map(ws, parse_rent_roll.COLUMNS)
# ws.cell(..., value=None) is a no-op in openpyxl — None means "no value
# supplied" — so clear through the attribute instead.
for c in range(1, ws.max_column + 1):
    for r in range(hdr - 1, hdr + 1):
        ws.cell(row=r, column=c).value = None
wb.save(bad)
raises("missing header is refused", lambda: parse_rent_roll.parse(bad), needle="header")

# positive: columns reordered and one inserted must still parse
shifted = copy(RR)
wb = openpyxl.load_workbook(shifted)
ws = wb["Rent Roll"]
ws.insert_cols(4)
ws.cell(row=parse_rent_roll.header_map(ws, parse_rent_roll.COLUMNS)[1], column=4,
        value="Some New Column")
wb.save(shifted)
try:
    rr2 = parse_rent_roll.parse(shifted)
    ok("survives an inserted column",
       rr2["totals"]["units"] == 263 and rr2["totals"]["market_rent"] == 1701103.0,
       f'{rr2["totals"]["units"]} units, market {rr2["totals"]["market_rent"]:,}')
except Exception as e:  # noqa: BLE001
    ok("survives an inserted column", False, str(e)[:80])

# ------------------------------------------------------------- delinquency
DQ = os.path.join(FIX, "delinquency.xlsx")
print("\ndelinquency parser")
dq = parse_delinquency.parse(DQ)
s = dq["summary"]
ok("property identified", dq["property_code"] == "p0005611",
   f'{dq["property"]} / {dq["property_code"]}')
ok("units with a balance matches", s["units_with_balance"] == 34, str(s["units_with_balance"]))
ok("gross owed matches", s["gross_owed"] == 57197.52, f'{s["gross_owed"]:,}')
ok("credits match", s["credits"] == -13750.48, f'{s["credits"]:,}')
ok("net matches", s["net"] == 43447.04, f'{s["net"]:,}')
ok("aging buckets match",
   [s["aging"][k] for k in ("d0_30", "d31_60", "d61_90", "over90")]
   == [46364.27, 6708.29, 539.13, 3585.83], str(s["aging"]))
ok("retail split out", (s["retail_balance"], s["retail_over90"]) == (5040.0, 4240.0),
   f'{s["retail_balance"]:,} of which {s["retail_over90"]:,} over 90 days')
ok("retail excluded from residential aging",
   all(r["residential"] for r in dq["residents"] if r["unit"] != "NONRES01"))
ok("total_all ties to Grand Total", s["total_all"] == 48487.04, f'{s["total_all"]:,}')
hard = [c for c in dq["checks"] if c["check"] != "as-of date found"]
ok("all tie-outs pass", all(c["ok"] for c in hard),
   ", ".join(c["check"] for c in hard if not c["ok"]) or "all")
ok("as-of absence is reported, not silent",
   any(c["check"] == "as-of date found" and not c["ok"] for c in dq["checks"])
   or dq["as_of"] is not None,
   f'as_of={dq["as_of"]} (this fixture omits the annotation column that carries it)')

# negative: aging no longer sums to the balance
bad = copy(DQ)
wb = openpyxl.load_workbook(bad)
ws = wb["Delinquency"]
fields, hdr = header_map(ws, parse_delinquency.COLUMNS)
ws.cell(row=hdr + 2, column=fields["d0_30"], value=99999)
wb.save(bad)
raises("aging that does not sum is refused", lambda: parse_delinquency.parse(bad),
       needle="aging")

# negative: Grand Total row removed -> nothing to tie out against
bad = copy(DQ)
wb = openpyxl.load_workbook(bad)
ws = wb["Delinquency"]
gt = parse_delinquency._grand_total(ws, header_map(ws, parse_delinquency.COLUMNS)[0])
ws.delete_rows(gt["row"] - 1, 2)          # both the property Total and Grand Total
wb.save(bad)
raises("missing Grand Total is refused", lambda: parse_delinquency.parse(bad))

# non-strict mode returns the problems instead of raising, for callers that
# want to log and continue
bad = copy(DQ)
wb = openpyxl.load_workbook(bad)
ws = wb["Delinquency"]
fields, hdr = header_map(ws, parse_delinquency.COLUMNS)
ws.cell(row=hdr + 2, column=fields["d0_30"], value=99999)
wb.save(bad)
loose = parse_delinquency.parse(bad, strict=False)
ok("strict=False reports instead of raising", bool(loose["problems"]),
   loose["problems"][0][:70] if loose["problems"] else "no problems recorded")

# ------------------------------------------- multi-property delinquency file
# Palma arrives as one export covering rspalman and rspalmas. The parser used to
# read only the first section and then compare it against the file's Grand Total,
# which spans both — a $150 gap that made it refuse a perfectly good report.
print("\nmulti-property delinquency")
MULTI = os.path.join(FIX, "delinquency_palma_multi.xlsx")
if not os.path.exists(MULTI):
    print("  SKIP  fixture missing — run scripts/make_multi_fixture.py " + MULTI)
else:
    m = parse_delinquency.parse(MULTI)          # strict: must not raise
    ok("both property codes found", m["property_codes"] == ["rspalman", "rspalmas"],
       str(m["property_codes"]))
    ok("top-level property left unset for a multi-property file",
       m["property"] is None and m["property_code"] is None,
       "so no caller can file a combined total under one of them")
    ok("one section per property", len(m["sections"]) == 2,
       f'{len(m["sections"])} sections')
    north = next(x for x in m["sections"] if x["property_code"] == "rspalman")
    south = next(x for x in m["sections"] if x["property_code"] == "rspalmas")
    ok("Palma North ties to its own Total row",
       north["summary"]["net"] == -20486.36, f'{north["summary"]["net"]:,.2f}')
    ok("Palma South ties to its own Total row",
       south["summary"]["net"] == -150.0, f'{south["summary"]["net"]:,.2f}')
    ok("North aging matches the report",
       [north["summary"]["aging"][k] for k in ("d0_30", "d31_60", "d61_90", "over90")]
       == [14989.88, 259.29, 10345.54, 11846.64], str(north["summary"]["aging"]))
    ok("combined net ties to the Grand Total", m["summary"]["net"] == -20636.36,
       f'{m["summary"]["net"]:,.2f}')
    ok("prepayment-heavy property still reports positive aging",
       m["summary"]["gross_owed"] == 37441.35 and m["summary"]["net"] < 0,
       f'gross {m["summary"]["gross_owed"]:,.2f} on a net credit of {m["summary"]["net"]:,.2f}')
    ok("every tie-out passes on the real layout", all(c["ok"] for c in m["checks"]),
       ", ".join(c["check"] for c in m["checks"] if not c["ok"]) or "all")

    # negative: break one section's own Total row. The Grand Total still agrees
    # with the sum of the rows, so only the per-section check can catch this.
    bad = copy(MULTI)
    wb = openpyxl.load_workbook(bad)
    ws = wb["Report1"]
    ws.cell(row=14, column=12, value=-99999)
    wb.save(bad)
    raises("a broken section Total is refused", lambda: parse_delinquency.parse(bad),
           needle="section rspalman")


# ---------------------------------------------------- PII must not persist
# The parsers read tenant names because the reports contain them. Nothing may
# write them to disk. This asserts the scrub rather than trusting it.
print("\nPII scrub")
import build_metrics  # noqa: E402

rr_names = {u["resident_name"] for u in rr["units"] if u.get("resident_name")}
ok("parser does read tenant names (so the scrub is doing real work)",
   len(rr_names) > 100, f"{len(rr_names)} names in the parse")

scrubbed = build_metrics.scrub(rr)
flat = str(scrubbed)
ok("scrub removes every tenant name",
   not any(n in flat for n in rr_names),
   f'{sum(1 for n in rr_names if n in flat)} names survived')
ok("scrub removes the PII keys",
   not any(k in flat for k in ("resident_name", "resident_code")))
ok("scrub keeps the numbers the dashboard needs",
   scrubbed["totals"]["market_rent"] == 1701103.0 and len(scrubbed["units"]) == 263,
   f'{len(scrubbed["units"])} units, market {scrubbed["totals"]["market_rent"]:,}')
ok("occupied flag survives the scrub (derived before it)",
   sum(1 for u in scrubbed["units"] if u.get("occupied")) > 200,
   f'{sum(1 for u in scrubbed["units"] if u.get("occupied"))} occupied')

dq_names = {r["resident_name"] for r in dq["residents"] if r.get("resident_name")}
dq_scrubbed = str(build_medium := build_metrics.scrub(dq))
ok("scrub removes delinquency names too",
   not any(n in dq_scrubbed for n in dq_names),
   f'{len(dq_names)} names in the parse, 0 survive')

# and end to end through the real writer
tmpdir = tempfile.mkdtemp()
_orig_data = build_metrics.DATA
try:
    import pathlib
    build_metrics.DATA = pathlib.Path(tmpdir)
    fp = build_metrics.store_rent_roll({"slug": "t"}, rr)
    written = open(fp).read()
    ok("store_rent_roll writes no names",
       not any(n in written for n in rr_names),
       os.path.basename(str(fp)))
finally:
    build_metrics.DATA = _orig_data
    shutil.rmtree(tmpdir, ignore_errors=True)

print(f"\n{len(PASS)} passed, {len(FAIL)} failed")
for n in FAIL:
    print("  failed: " + n)
sys.exit(1 if FAIL else 0)
