"""
Guard tests for parse_leasing_funnel and parse_concession_burnoff.

Self-contained: builds synthetic workbooks that reproduce the structures the
inspect workflow dumped from the real 2026-08-04 / 2026-08-18 funnel exports
and the 2026-08-10 burn-off export, then checks:

  - both parsers read their file and tie out
  - a funnel with two communities routes each to its own property
  - deliberately broken files are REFUSED (bad tie-out, moved header)
  - the burn-off's resident names never reach any stored file
  - the burn-off stores nowhere while the export names no property

Run: python scripts/test_funnel_and_concessions.py
"""

import json
import pathlib
import shutil
import sys
import tempfile

import openpyxl

sys.path.insert(0, str(pathlib.Path(__file__).parent))
import build_metrics as bm                      # noqa: E402
import parse_concession_burnoff as pcb          # noqa: E402
import parse_leasing_funnel as plf              # noqa: E402

FAKE_NAMES = ["Testfirst Testlastone", "Testfirst Testlasttwo"]

FUNNEL_HDR_PORT = ["Service Start Date", "Prospects Engaged", "Responses to Prospects",
                   "Follow Ups Sent", "Appointments Scheduled", "AI Scheduled Appointments",
                   "% of Appts Booked by AI", "Lead to Appointment Rate",
                   "Lead to Appointment Rate (excl. ApartmentLis", "Hours Saved by AI",
                   "After Hours Activity>>", "Prospects Engaged", "Responses to Prospects",
                   "Follow Ups Sent", "AI Scheduled Appointments", "Hours Saved by AI",
                   "Onsite Interaction>>", "Incorrect Agent Takeover",
                   "AI Handed Off (Handoff Rate)", "Lead to Show Rate (Total)"]
FUNNEL_HDR_COMM = (["Community"] + FUNNEL_HDR_PORT +
                   ["Lead to Application Rate", "Lead to Lease Rate",
                    "Number of Leases", "PropertyId"])


def make_funnel(path, break_tie_out=False):
    wb = openpyxl.Workbook()
    # two communities; the portfolio is their sum
    comm = [
        ["335 3rd Street", "2026-01-01", 100, 800, 200, 40, 30, .75, .40, .40, 60.0,
         None, 50, 300, 80, 5, 25.0, None, 0, .20, .10, .02, 0, 0, 938531],
        ["The Landing", "2026-01-01", 62, 479, 171, 28, 19, .68, .45, .45, 46.58,
         None, 32, 193, 46, 3, 16.08, None, 0, .15, .12, .04, .01, 3, 938600],
    ]
    port = ["2026-01-01"]
    for i in range(2, 21):                      # sum the count columns
        a, b = comm[0][i], comm[1][i]
        port.append(None if a is None and b is None
                    else (a or 0) + (b or 0) if isinstance(a, (int, float)) or isinstance(b, (int, float))
                    else None)
    # rates in the portfolio row are not sums; put plausible numbers back
    for idx, v in ((6, .72), (7, .42), (8, .42), (18, .18), (19, .11)):
        port[idx] = v
    if break_tie_out:
        port[1] = 999                           # prospects engaged no longer ties out

    ws = wb.active
    ws.title = "Portfolio to Date"
    ws.append(FUNNEL_HDR_PORT)
    ws.append(port)
    ws = wb.create_sheet("Portfolio by Month")
    ws.append(["Month"] + FUNNEL_HDR_PORT[1:])
    ws.append(["2026-08", 20, 200, 60, 8, 5, .63, .40, .40, 15.0,
               None, 10, 70, 20, 1, 6.0, None, 0, .2, .1])
    ws = wb.create_sheet("Portfolio by Week")
    ws.append(["Week Of"] + FUNNEL_HDR_PORT[1:17])
    ws.append(["2026-08-10", 10, 90, 30, 4, 2, .5, .4, .4, 8.0, None, 5, 30, 9, 0, 3.0])
    ws = wb.create_sheet("Comm to Date")
    ws.append(FUNNEL_HDR_COMM)
    for row in comm:
        ws.append(row)
    ws = wb.create_sheet("Comm by Month")
    ws.append(["Community", "Month"] + FUNNEL_HDR_COMM[2:])
    ws.append(["335 3rd Street", "2026-08", 15, 150, 45, 6, 4, .67, .4, .4, 11.0,
               None, 8, 50, 15, 1, 4.5, None, 0, .2, .1, .02, 0, 0, 938531])
    ws.append(["The Landing", "2026-08", 5, 50, 15, 2, 1, .5, .4, .4, 4.0,
               None, 2, 20, 5, 0, 1.5, None, 0, .1, .1, .04, .01, 1, 938600])
    for name in ("Comm by Month, Channel", "Comm by Mth,Chan,Source"):
        wb.create_sheet(name).append(["Community", "Month", "Channel"])
    wb.save(path)


def make_burnoff(path, break_tie_out=False, move_header=False):
    """Deliberately gritty, because the first REAL export broke a clean-fixture
    parser: money as text ("(1,500.00)", "$3,600", "1,234.00"), a section-text
    row, an MTM lease term, and a dash where a number would be."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report1"
    ws.append(["Concession Burn Off"])
    ws.append(["For Selected Properties"])
    ws.append(["As Of = 08/10/2026"])
    ws.append(["Header moved" if move_header else "Unit", "Unit", "Resident", "Name",
               "Move In", "Lease Start", "Total", "Current Lease", "Current Lease",
               "Concession", "Lease", "Market", "Lease", "Current"])
    ws.append([None, "Type", None, None, "Date", "Date", "Recurring",
               "Concessions", "Concessions", "End Date", "Term", "Rent", "Rent", "Month"])
    ws.append([None, None, None, None, None, None, "Concessions", None, "Remaining"])
    # section 1: labeled with a property alias, mixed numeric/text money
    ws.append(["The Landing"])                     # heading row -> section label
    ws.append(["101", "A1", FAKE_NAMES[0].split()[0], FAKE_NAMES[0].split()[1],
               "2026-05-01", "2026-05-01", -1500.0, -1500.0, -1000.0,
               "2027-04-30", 12, 2800.0, 2650.0, 0])
    ws.append(["202", "B2", FAKE_NAMES[1].split()[0], FAKE_NAMES[1].split()[1],
               "2026-06-15", "2026-06-15", "(2,400.00)", "(2,400.00)", "(2,200.00)",
               "2027-06-14", "MTM", "$3,600", "3,400.00", "-"])
    ws.append(["303", "C3", None, None, None, None,   # unit label, no figures
               None, None, None, None, None, "-", "-", "-"])
    ws.append([None, None, None, None, None, None, -3900.0,   # subtotal
               -3900.0, -3200.0, None, None, 6400.0, 6050.0, 0])
    # section 2: another property block
    ws.append(["Palma"])
    ws.append(["B12", "S1", FAKE_NAMES[0].split()[0], FAKE_NAMES[0].split()[1],
               "2026-07-01", "2026-07-01", -1833.0, -1833.0, -1833.0,
               "2027-06-30", 12, 2500.0, 2450.0, 0])
    ws.append([None, None, None, None, None, None, -1833.0,   # subtotal
               -1833.0, -1833.0, None, None, 2500.0, 2450.0, 0])
    # grand total across both sections
    total_g = -5733.0 + (100 if break_tie_out else 0)
    ws.append([None, None, None, None, None, None, total_g,
               -5733.0, -5033.0, None, None, 8900.0, 8500.0, 0])
    wb.save(path)


def main():
    tmp = pathlib.Path(tempfile.mkdtemp())
    passed = failed = 0

    def check(label, ok, detail=""):
        nonlocal passed, failed
        passed += ok
        failed += (not ok)
        print(f"  {'PASS' if ok else 'FAIL'}  {label}{('  ' + detail) if detail and not ok else ''}")

    print("1. leasing funnel — clean export")
    fpath = tmp / "leasing_funnel_report_2026-08-18.xlsx"
    make_funnel(fpath)
    p = plf.parse(str(fpath))
    check("as_of from filename", p["as_of"] == "2026-08-18", p["as_of"])
    check("two communities", len(p["sections"]) == 2)
    check("tie-outs all ok", all(c["ok"] for c in p["checks"]))
    s335 = next(s for s in p["sections"] if s["property_code"] == "335 3rd Street")
    check("prospects engaged", s335["to_date"]["prospects_engaged"] == 100)
    check("after-hours prefix distinct",
          s335["to_date"].get("after_hours_activity_prospects_engaged") == 50)
    check("leases carried", s335["to_date"]["number_of_leases"] == 0)

    print("2. leasing funnel — broken portfolio tie-out is refused")
    bad = tmp / "leasing_funnel_report_2026-08-19.xlsx"
    make_funnel(bad, break_tie_out=True)
    try:
        plf.parse(str(bad))
        check("refused", False, "parse succeeded on a broken file")
    except ValueError:
        check("refused", True)

    print("3. funnel routes per community through aliases")
    bm.DATA = tmp / "data"
    props, code_to_prop = bm.load_properties()
    check("alias resolves", code_to_prop.get("335 3rd street", {}).get("slug") == "335-third-street")
    for sec in p["sections"]:
        prop = code_to_prop.get(sec["property_code"].lower())
        check(f"routes {sec['property_code']!r}", prop is not None)
        if prop:
            one = dict(p)
            one["sections"] = [sec]
            bm.store_leasing_funnel(prop, one)
    stored = json.loads((bm.DATA / "335-third-street" / "leasing_funnel.json").read_text())
    check("stored community fields", stored["community"] == "335 3rd Street"
          and stored["to_date"]["prospects_engaged"] == 100
          and len(stored["by_month"]) == 1)

    print("4. concession burn-off — clean export")
    cpath = tmp / "ConcessionBurnOff08_10_2026.xlsx"
    make_burnoff(cpath)
    c = pcb.parse(str(cpath))
    check("as_of from A3", c["as_of"] == "2026-08-10", c["as_of"])
    check("sections are labeled, so not unattributed", c["unattributed"] is False)
    check("unit count excludes total, section and empty rows", c["unit_count"] == 3)
    check("two sections with their heading labels",
          [s["label"] for s in c["sections"]] == ["The Landing", "Palma"])
    all_units = [u for s in c["sections"] for u in s["units"]]
    u202 = next(u for u in all_units if u["unit"] == "202")
    check("text money parsed", u202["recurring_concessions"] == -2400.0
          and u202["market_rent"] == 3600.0 and u202["lease_term"] == "MTM")
    check("per-section and grand tie-outs all ok", all(k["ok"] for k in c["checks"]))
    check("section labels route through aliases",
          all(code_to_prop.get(s["label"].lower()) for s in c["sections"]))
    one = dict(c); one["sections"] = [c["sections"][0]]
    bm.store_concessions(code_to_prop["the landing"], one)
    landed = json.loads((bm.DATA / "the-landing" / "concessions.json").read_text())
    check("stored aggregate only, no unit rows",
          "units" not in landed and landed["unit_count"] == 2
          and landed["totals"]["recurring_concessions"] == -3900.0)
    blob2 = json.dumps(landed)
    check("no resident name in the stored file",
          all(n.split()[0] not in blob2 for n in FAKE_NAMES))
    check("totals tie out", all(k["ok"] for k in c["checks"]) and
          c["totals"]["recurring_concessions"] == -5733.0)
    blob = json.dumps({k: v for k, v in c.items() if k != "sections"})
    check("no resident name in the parse output",
          all(n.split()[0] not in blob and n.split()[1] not in blob for n in FAKE_NAMES))

    print("5. concession burn-off — refusals")
    bad = tmp / "ConcessionBurnOff_bad.xlsx"
    make_burnoff(bad, break_tie_out=True)
    try:
        pcb.parse(str(bad))
        check("bad tie-out refused", False, "parse succeeded")
    except ValueError:
        check("bad tie-out refused", True)
    moved = tmp / "ConcessionBurnOff_moved.xlsx"
    make_burnoff(moved, move_header=True)
    try:
        pcb.parse(str(moved))
        check("moved header refused", False, "parse succeeded")
    except ValueError:
        check("moved header refused", True)

    print("6. stored funnel file carries no person-shaped keys")
    hits = [k for k in stored if any(w in k for w in ("resident", "tenant", "name"))]
    check("no person-shaped keys", not hits, str(hits))

    shutil.rmtree(tmp)
    print(f"\n{passed} passed, {failed} failed")
    sys.exit(1 if failed else 0)


if __name__ == "__main__":
    main()
