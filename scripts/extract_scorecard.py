#!/usr/bin/env python3
"""Build docs/scorecard.json — the KPI health scorecard.

Reads the "KPI (Flipped Axis)" sheet of the KPI scorecard workbook, which
defines the metric groups in its merged header row, and (since v10) the
"KPI Target Ranges" sheet, which publishes the numeric bands behind each
status plus the lease-up overrides for Palma North/South.

Statuses are still the literal symbols the workbook holds (set by hand);
the ranges are published alongside them under "thresholds" so the dashboard
can show what each symbol means, not yet to recompute the symbols.

Usage: python scripts/extract_scorecard.py <path-to-KPI_Scorecard.xlsx>
"""
import datetime
import json
import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter

SRC = sys.argv[1] if len(sys.argv) > 1 else "KPI_Scorecard_Formatted_v10.xlsx"
OUT = "docs/scorecard.json"
SHEET = "KPI (Flipped Axis)"
RANGES_SHEET = "KPI Target Ranges"

HEADER_GROUP_ROW = 7
HEADER_METRIC_ROW = 8
FIRST_DATA_ROW = 9
FIRST_METRIC_COL = 3          # C

# From the workbook's own legend cells (C4/G4/K4) — do not re-invent these.
# v10 changed the in-range band to white: it carries no color indicator, only
# the ● symbol. The dashboard renders it clear/neutral for the same reason.
LEGEND = [
    {"symbol": "▲", "state": "exceeding", "label": "Exceeding KPI target range",
     "color": "green", "xlsx_fill": "FF00B050"},
    {"symbol": "●", "state": "in_range", "label": "In KPI target range",
     "color": "none", "xlsx_fill": "FFFFFFFF"},
    {"symbol": "▼", "state": "below", "label": "Below KPI target range",
     "color": "red", "xlsx_fill": "FFFF0000"},
]
STATE_BY_SYMBOL = {l["symbol"]: l["state"] for l in LEGEND}

# KPIs that report a figure rather than a grade. A distribution has no single
# direction it can be good or bad in — the ranges sheet says as much in its own
# basis note for this one ("A distribution cannot be scored") — so the cell
# carries its numbers and no status at all: no symbol, no colour, and it is
# excluded from the at-or-above-target counts rather than inflating them.
UNSCORED = {"Split Between 30/60/90"}

# KPIs on the workbook's grid that the dashboard does not report at all. Dropped
# here rather than hidden on the page, so nothing downstream carries a column
# with no home: not the matrix, not a property's row, not the thresholds table,
# not the coverage counts. The workbook keeps its own column either way.
OMITTED_METRICS = {"# of offers that are 30 days"}

# Rows on the workbook's grid that are not published. The grid is the analyst's
# working sheet and carries properties the dashboard does not report on; each
# one published adds a column of hand-set symbols to the matrix and another
# property to the portfolio roll-up, with no measured value behind any of it.
EXCLUDED_PROPERTIES = {"Fitzgerald", "2177 Third"}

# Scorecard property label -> slug in config/properties.json. Any slug named
# here that is missing from the master is reported below and downgraded to
# None, so this map cannot silently drift out of sync with the config. A
# property that maps to None still renders, just without a link to a property
# view that does not exist.
SLUGS = {
    "Chorus": "chorus",
    "Landing": "the-landing",
    "335 Third": "335-third-street",
    "Madelon": "madelon",
    "Palma": "palma",
}

try:
    KNOWN_SLUGS = {p["slug"] for p in json.load(open("config/properties.json"))["properties"]}
except OSError:
    KNOWN_SLUGS = None                      # run from elsewhere; skip the check

STALE = sorted(s for s in SLUGS.values()
               if s and KNOWN_SLUGS is not None and s not in KNOWN_SLUGS)
for label, slug in SLUGS.items():
    if slug in STALE:
        SLUGS[label] = None

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb[SHEET]

# ---- metrics, in sheet order, tagged with the group that spans them ----
metrics, group = [], None
for col in range(FIRST_METRIC_COL, ws.max_column + 1):
    g = ws.cell(row=HEADER_GROUP_ROW, column=col).value
    if g:
        group = str(g).strip()
    name = ws.cell(row=HEADER_METRIC_ROW, column=col).value
    if not name:
        continue
    metrics.append({"name": str(name).strip(), "group": group,
                    "col": get_column_letter(col)})

dropped = sorted({m["name"] for m in metrics} & OMITTED_METRICS)
metrics = [m for m in metrics if m["name"] not in OMITTED_METRICS]
for n in dropped:
    print(f"[skip] {n} \u2014 not published (OMITTED_METRICS)")
gone = sorted(OMITTED_METRICS - set(dropped))
if gone:
    print(f"  WARNING: OMITTED_METRICS names {', '.join(gone)}, which the grid "
          f"does not have \u2014 a renamed column would silently start publishing")

# ---- one record per property ----
properties = []
for row in range(FIRST_DATA_ROW, ws.max_row + 1):
    label = ws.cell(row=row, column=2).value
    if not label:
        continue
    label = str(label).strip()
    if label in EXCLUDED_PROPERTIES:
        print(f"[skip] {label} — excluded from the dashboard (EXCLUDED_PROPERTIES)")
        continue
    statuses = {}
    for i, m in enumerate(metrics):
        sym = ws.cell(row=row, column=FIRST_METRIC_COL + i).value
        if m["name"] in UNSCORED:
            state = None            # reported, not graded — see UNSCORED above
        else:
            state = STATE_BY_SYMBOL.get(str(sym).strip()) if sym else None
        statuses[m["name"]] = state
    properties.append({
        "label": label,
        "slug": SLUGS.get(label),
        "statuses": statuses,
        # Measured values behind each status. The workbook holds only the
        # symbol today, so these are empty and the dashboard reserves the
        # space and shows a placeholder. Fill "raw" with the number and
        # "display" with the string to print ("94.2%", "$1,240", "12").
        "values": {m["name"]: {"raw": None, "display": None} for m in metrics},
        # counts, scored, at_or_above, below_metrics and coverage are all filled
        # by recompute below, from the cells that carry a measurement. Straight
        # out of this script that is none of them, which is the honest answer:
        # the workbook's symbols are hand-set, and until a report supplies a
        # number there is nothing to grade.
    })

# ---- target ranges (new in v10) ----------------------------------------
# The ranges sheet spells each KPI slightly differently in two places
# ("Total Deliquency" on the grid, "Total Delinquency" here; "# of work
# orders" vs "# of open work orders"). Thresholds are keyed by the GRID's
# metric names, since that is what the dashboard renders.
def _norm(s):
    return re.sub(r"\s+", " ", str(s)).strip().casefold()

ALIASES = {                                   # ranges-sheet name -> grid name
    "total delinquency": "Total Deliquency",
    "# of open work orders": "# of work orders",
    "open eliseai tasks": "Open Elise Tasks",
}
GRID_BY_NORM = {_norm(m["name"]): m["name"] for m in metrics}

thresholds, leaseup_overrides = {}, []
if RANGES_SHEET in wb.sheetnames:
    rs = wb[RANGES_SHEET]

    def find(label, col, after=1):
        for r in range(after, rs.max_row + 1):
            if _norm(rs.cell(row=r, column=col).value or "") == _norm(label):
                return r
        raise SystemExit(f"FATAL: no '{label}' row on {RANGES_SHEET!r} — layout moved")

    hdr = find("Category", 2)
    category = None
    for r in range(hdr + 1, rs.max_row + 1):
        kpi = rs.cell(row=r, column=3).value
        if not kpi:
            if rs.cell(row=r, column=2).value:      # a section below the table
                break
            continue
        category = str(rs.cell(row=r, column=2).value or category or "").strip() or category
        name = ALIASES.get(_norm(kpi)) or GRID_BY_NORM.get(_norm(kpi)) or str(kpi).strip()
        thresholds[name] = {
            "group": category,
            "how": rs.cell(row=r, column=4).value,
            "direction": rs.cell(row=r, column=5).value,
            "exceeding": rs.cell(row=r, column=6).value,
            "in_range": rs.cell(row=r, column=7).value,
            "below": rs.cell(row=r, column=8).value,
            "green_cutoff": rs.cell(row=r, column=9).value,
            "red_cutoff": rs.cell(row=r, column=10).value,
            "basis": rs.cell(row=r, column=11).value,
        }

    ov_hdr = find("KPI", 3, after=find("Lease-up overrides", 2))
    for r in range(ov_hdr + 1, rs.max_row + 1):
        kpi = rs.cell(row=r, column=3).value
        if not kpi:
            break
        leaseup_overrides.append({
            "kpi": ALIASES.get(_norm(kpi)) or GRID_BY_NORM.get(_norm(kpi)) or str(kpi).strip(),
            "approach": rs.cell(row=r, column=4).value,
            "exceeding": rs.cell(row=r, column=5).value,
            "in_range": rs.cell(row=r, column=6).value,
            "below": rs.cell(row=r, column=7).value,
        })

    # an omitted KPI's range goes with it, so the table cannot describe a cell
    # that is no longer anywhere on the dashboard
    for n in OMITTED_METRICS & set(thresholds):
        del thresholds[n]

    unmatched = sorted(set(thresholds) - {m["name"] for m in metrics})
    if unmatched:
        print("  WARNING: ranges for KPIs not on the grid: " + ", ".join(unmatched))
    missing = sorted({m["name"] for m in metrics} - set(thresholds))
    if missing:
        print("  WARNING: grid KPIs with no published range: " + ", ".join(missing))

# ---- legend drift guard --------------------------------------------------
# The LEGEND above mirrors the workbook's legend cells; if the workbook
# changes its colors again, fail loudly rather than publish stale semantics.
for cell_ref, expect in (("C4", "FF00B050"), ("G4", "FFFFFFFF"), ("K4", "FFFF0000")):
    f = ws[cell_ref].fill
    got = f.fgColor.rgb if f.patternType else None
    if got != expect:
        raise SystemExit(f"FATAL: legend cell {cell_ref} fill is {got}, expected {expect} — "
                         "the workbook legend changed; update LEGEND to match")

data = {
    "meta": {
        "generated_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "source_workbook": SRC.split("/")[-1],
        "source_sheet": SHEET,
        "note": ("Statuses are set by hand in the workbook. The numeric bands behind "
                 "each status are published under \"thresholds\"; in-range carries no "
                 "color indicator, by design. Palma North/South are in lease-up and "
                 "four of their KPIs are scored against the lease-up overrides. "
                 "A cell carries a colour and counts toward the tally only where a "
                 "report has supplied a number and the published band could grade "
                 "it; the workbook's hand-set symbol is kept in \"statuses\" and in "
                 "\"status_workbook\" but is not treated as a result."),
    },
    "legend": LEGEND,
    # reported, not graded: no status, no colour, not counted in "scored"
    "unscored": sorted(UNSCORED & {m["name"] for m in metrics}),
    "groups": [{"name": g, "metrics": [m["name"] for m in metrics if m["group"] == g]}
               for g in dict.fromkeys(m["group"] for m in metrics)],
    "metrics": [{"name": m["name"], "group": m["group"]} for m in metrics],
    "properties": properties,
    # filled by recompute below, off the same definition the populate scripts use
    "portfolio": {},
    # Keyed by the grid's metric names. Each entry: group, how (measurement),
    # direction, the three display bands, the numeric green/red cutoffs, and
    # the sourcing rationale.
    "thresholds": thresholds or None,
    # Palma North/South are in lease-up; these KPIs are scored against
    # different bands there (see each entry's "approach").
    "leaseup_overrides": leaseup_overrides or None,
}

# One definition of what counts, shared with the populate scripts rather than
# repeated here — see populate_scorecard.graded.
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from populate_scorecard import recompute
recompute(data)

with open(OUT, "w") as f:
    json.dump(data, f, separators=(",", ":"))

pf = data["portfolio"]
cov = pf["coverage"]
print(f"wrote {OUT} ({os.path.getsize(OUT)} bytes)")
print("  NOTE: every measured value is now null. Re-run "
      "scripts/populate_scorecard.py to put the measured numbers back.")
print(f"  {len(properties)} properties x {len(metrics)} metrics = {cov['total']} cells")
print(f"  graded {cov['graded']}  ·  reported not graded {cov['reported_ungraded']}"
      f"  ·  awaiting a feed {cov['awaiting']}")
print("  " + "  ".join(f"{k}={v}" for k, v in pf["counts"].items()))
unmapped = [p["label"] for p in properties if not p["slug"]]
if unmapped:
    print("  no property view / not in the master: " + ", ".join(unmapped))
if STALE:
    print("  WARNING: slugs mapped here but missing from config/properties.json: "
          + ", ".join(STALE))
