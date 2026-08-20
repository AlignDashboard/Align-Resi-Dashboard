"""
parse_leasing_funnel.py
-----------------------
Parses EliseAI's weekly leasing funnel export (leasing_funnel_report_YYYY-MM-DD
.xlsx). This is the weekly baseline feed in the owner's two-feed design: the
daily "Leasing AI Daily Report" emails update it between exports.

Seven sheets, verified against the 2026-08-04 and 2026-08-18 exports via the
inspect workflow:

  Portfolio to Date / by Month / by Week   whole-portfolio roll-ups
  Comm to Date / by Month                  the same figures per community
  Comm by Month, Channel / by Mth,Chan,Source   channel splits (not parsed --
                                           marketing detail, nothing the
                                           dashboard reports on)

The file carries NO person-level data (confirmed by the inspector's name scan
on both exports), only counts and rates. Communities are named in the export's
own labels ("335 3rd Street"), which config/properties.json maps to a property
via each property's "aliases" -- the same labels the building-metrics CSV uses.

Tie-out: every count column on the portfolio to-date sheet must equal the sum
of the communities' to-date rows. A file that fails is refused rather than
stored, matching the delinquency parser's behaviour.
"""

import datetime
import os
import re

import openpyxl

# Sheets read. The two channel sheets are deliberately not parsed.
SHEET_PORT_TODATE = "Portfolio to Date"
SHEET_PORT_MONTH = "Portfolio by Month"
SHEET_PORT_WEEK = "Portfolio by Week"
SHEET_COMM_TODATE = "Comm to Date"
SHEET_COMM_MONTH = "Comm by Month"

# Count columns that must tie out between the portfolio sheet and the sum of
# the community sheets. Rates are ratios and do not sum.
TIE_OUT_FIELDS = ("prospects_engaged", "responses_to_prospects",
                  "follow_ups_sent", "appointments_scheduled",
                  "ai_scheduled_appointments", "number_of_leases")


def _key(header, prefix=""):
    """'Lead to Appointment Rate (excl. ApartmentLis' -> lead_to_appointment_rate_excl_apartmentlist"""
    k = re.sub(r"[^a-z0-9]+", "_", str(header).lower()).strip("_")
    # the excl-ApartmentList header arrives truncated by the export itself
    k = k.replace("excl_apartmentlis", "excl_apartmentlist")
    return prefix + k


def _headers(ws):
    """Column keys from row 1. The sheet repeats header names after the
    'After Hours Activity>>' marker (the after-hours block restates Prospects
    Engaged, Responses, Follow Ups...), so a name already seen gets the current
    marker's prefix; a unique name stays plain even after a marker, because
    Number of Leases and PropertyId sit after 'Onsite Interaction>>' and must
    keep their own names. Marker columns carry no data and map to None."""
    keys, seen, prefix = [], set(), ""
    for cell in ws[1]:
        v = cell.value
        if v is None:
            keys.append(None)
            continue
        s = str(v).strip()
        if s.endswith(">>"):
            prefix = _key(s.rstrip(">").strip()) + "_"
            keys.append(None)
            continue
        k = _key(s)
        if k in seen:
            k = prefix + k if prefix and (prefix + k) not in seen else k + "_2"
        seen.add(k)
        keys.append(k)
    return keys


def _rows(ws):
    keys = _headers(ws)
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        row = {}
        for k, v in zip(keys, r):
            if k is None:
                continue
            if isinstance(v, datetime.datetime):
                v = v.date().isoformat()
            elif isinstance(v, datetime.date):
                v = v.isoformat()
            row[k] = v
        out.append(row)
    return out


def _as_of(path):
    """leasing_funnel_report_2026-08-18.xlsx -> 2026-08-18 (the export date)."""
    m = re.search(r"(20\d{2})-(\d{2})-(\d{2})", os.path.basename(path))
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else None


def parse(path, strict=True):
    wb = openpyxl.load_workbook(path, data_only=True)
    for name in (SHEET_PORT_TODATE, SHEET_COMM_TODATE, SHEET_COMM_MONTH):
        if name not in wb.sheetnames:
            raise ValueError(f"not a leasing funnel export: no {name!r} sheet "
                             f"(has {wb.sheetnames})")

    port_todate = _rows(wb[SHEET_PORT_TODATE])[0]
    port_month = _rows(wb[SHEET_PORT_MONTH]) if SHEET_PORT_MONTH in wb.sheetnames else []
    port_week = _rows(wb[SHEET_PORT_WEEK]) if SHEET_PORT_WEEK in wb.sheetnames else []
    comm_todate = _rows(wb[SHEET_COMM_TODATE])
    comm_month = _rows(wb[SHEET_COMM_MONTH])

    # one section per community, in the export's own label
    sections, checks = [], []
    for c in comm_todate:
        name = c.get("community")
        months = [m for m in comm_month if m.get("community") == name]
        months.sort(key=lambda m: str(m.get("month") or ""), reverse=True)
        sections.append({
            # the community label doubles as the routing key; properties.json
            # maps it to a slug via "aliases"
            "property_code": name,
            "community": name,
            "property_id": c.get("propertyid"),
            "service_start": c.get("service_start_date"),
            "to_date": {k: v for k, v in c.items()
                        if k not in ("community", "propertyid")},
            "by_month": [{k: v for k, v in m.items()
                          if k not in ("community", "propertyid")} for m in months],
        })

    # portfolio counts must equal the sum over communities
    for f in TIE_OUT_FIELDS:
        if f not in port_todate:
            continue
        port = port_todate.get(f) or 0
        comm = sum(c["to_date"].get(f) or 0 for c in sections)
        ok = abs(port - comm) < 0.5
        checks.append({"field": f, "portfolio": port, "communities": comm, "ok": ok})
        if strict and not ok:
            raise ValueError(f"tie-out failed: portfolio {f} = {port} but the "
                             f"communities sum to {comm} -- refusing to store")

    return {
        "report_type": "leasing_funnel",
        "as_of": _as_of(path),
        "source_file": os.path.basename(path),
        "portfolio": {"to_date": port_todate, "by_month": port_month,
                      "by_week": port_week},
        "sections": sections,
        "checks": checks,
    }


if __name__ == "__main__":
    import json
    import sys
    print(json.dumps(parse(sys.argv[1]), indent=2, default=str))
