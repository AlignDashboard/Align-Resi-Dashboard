"""
parse_delinquency.py
--------------------
Parses a Yardi rs_rp_DelinquencySummaryReport export into per-resident balances
plus the aging summary the dashboard shows.

Anchors on header labels, not positions, and ties the parsed rows out against
the report's own Grand Total.

The one judgement this encodes, taken from how the workbook already treats it:
the ground-floor retail tenant (unit code NONRES*) is reported alongside the
residents but is a single commercial receivable. Mixing it into resident aging
overstates resident credit risk, so it is split out. Residential gross is the
sum of positive balances and credits are the negatives, kept separate rather
than netted, because a prepayment is not a collection.

Usage:
    from parse_delinquency import parse
    result = parse("path/to/delinquency.xlsx")
"""
import json
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from xlsx_anchors import (LayoutError, header_map, norm,  # noqa: E402
                          rows_until)

COLUMNS = {
    "unit": r"(property\s*)?unit",
    "resident_code": r"resident code",
    "resident_name": r"resident (last )?name",
    "status": r"resident status",
    "total_charges": r"total charges",
    "future_charges": r"future charges",
    "d0_30": r"0\s*-\s*30",
    "d31_60": r"31\s*-\s*60",
    "d61_90": r"61\s*-\s*90",
    "over90": r"over\s*90",
    "prepayments": r"prepayment",
    "total_owed": r"total owed",
}

STOP = [r"^total\b", r"^grand total", r"^report total"]
NONRES = re.compile(r"^nonres", re.I)
AGING = ("d0_30", "d31_60", "d61_90", "over90")

# One export can cover several properties — Palma comes through as rspalman and
# rspalmas in a single file, each with its own group header and its own Total
# row, followed by one Grand Total for the file. A group header is
# "<code> - <name>"; its closing row is "Total <code> - <name>".
GROUP_HDR = re.compile(r"^([A-Za-z0-9._-]{3,20})\s*[-–]\s*(.+)$")
GRAND = re.compile(r"^grand total\b|^report total\b", re.I)
SECTION_TOTAL = re.compile(r"^total\s+(.+)$", re.I)


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").replace("$", "").strip()
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _property(ws, unit_col):
    """'p0005611 - The Landing' appears as a group header above the rows."""
    for r in range(1, min(ws.max_row, 40) + 1):
        v = ws.cell(row=r, column=unit_col).value
        if not isinstance(v, str):
            continue
        m = re.match(r"^([A-Za-z0-9._-]{4,20})\s*[-–]\s*(.+)$", v.strip())
        if m and not m.group(2).lower().startswith("the landing unit"):
            return m.group(2).strip(), m.group(1).strip()
    return None, None


def _as_of(ws):
    months = ["january", "february", "march", "april", "may", "june", "july",
              "august", "september", "october", "november", "december"]
    for r in range(1, min(ws.max_row, 40) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            m = re.search(r"as of\s*:?\s*(\d{1,2})\s+([A-Za-z]{3,9})\.?\s+(\d{4})", v, re.I)
            if m:
                mon = next((i for i, n in enumerate(months, 1)
                            if n.startswith(m.group(2).lower()[:3])), None)
                if mon:
                    return f"{m.group(3)}-{mon:02d}-{int(m.group(1)):02d}"
            m = re.search(r"as of\s*:?\s*(\d{1,2})/(\d{1,2})/(\d{4})", v, re.I)
            if m:
                return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return None


def _grand_total(ws, fields):
    for r in range(ws.max_row, 1, -1):
        label = norm(ws.cell(row=r, column=fields["unit"]).value)
        if re.search(r"grand total|^total\b", label):
            owed = _num(ws.cell(row=r, column=fields["total_owed"]).value)
            if owed is not None:
                return {"row": r, "total_owed": owed,
                        **{k: _num(ws.cell(row=r, column=fields[k]).value) or 0.0
                           for k in AGING}}
    return None


def _pick_sheet(wb):
    for name in wb.sheetnames:
        if "delinq" in name.lower():
            return wb[name]
    return wb[wb.sheetnames[0]]


def _summarise(residents):
    """The aging summary for one set of resident rows."""
    res = [r for r in residents if r["residential"]]
    ret = [r for r in residents if not r["residential"]]
    owing = [r for r in res if r["total_owed"] > 0]

    gross = round(sum(r["total_owed"] for r in owing), 2)
    credits = round(sum(r["total_owed"] for r in res if r["total_owed"] < 0), 2)
    # Aging is reported for accounts that owe; a credit balance has no bucket.
    aging = {k: round(sum(r[k] for r in owing), 2) for k in AGING}
    return {
        "units_with_balance": len(owing),
        "gross_owed": gross,
        "credits": credits,
        "net": round(gross + credits, 2),
        "aging": aging,
        "retail_balance": round(sum(r["total_owed"] for r in ret), 2),
        "retail_over90": round(sum(r["over90"] for r in ret), 2),
        "total_all": round(sum(r["total_owed"] for r in residents), 2),
        "resident_rows": len(res),
        "retail_rows": len(ret),
    }


def _read_sections(ws, fields, header_row):
    """Split the sheet into per-property sections.

    Returns (sections, grand_row). Each section is
    {code, property, residents, total_row, reported_total}. A file with a single
    property yields one section, so this reads the same as it always did.
    """
    unit_col = fields["unit"]
    sections, grand_row = [], None
    cur = None
    for r in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(row=r, column=unit_col).value
        label = str(raw).strip() if raw is not None else ""
        if not label:
            continue
        if GRAND.match(label):
            grand_row = {"row": r, "total_owed": _num(ws.cell(row=r, column=fields["total_owed"]).value)}
            continue
        st = SECTION_TOTAL.match(label)
        if st:
            # closing row for the section that is open
            if cur is not None:
                cur["total_row"] = r
                cur["reported_total"] = _num(ws.cell(row=r, column=fields["total_owed"]).value)
            continue
        gh = GROUP_HDR.match(label)
        # a group header has no money on its row; a unit row does
        if gh and _num(ws.cell(row=r, column=fields["total_owed"]).value) is None:
            cur = {"code": gh.group(1).strip(), "property": gh.group(2).strip(),
                   "residents": [], "total_row": None, "reported_total": None}
            sections.append(cur)
            continue
        # a data row: needs a resident code or a number to count
        if (ws.cell(row=r, column=fields.get("resident_code", unit_col)).value in (None, "")
                and _num(ws.cell(row=r, column=fields["total_owed"]).value) is None):
            continue
        if cur is None:                      # no group header seen (older exports)
            cur = {"code": None, "property": None, "residents": [],
                   "total_row": None, "reported_total": None}
            sections.append(cur)
        rec = {f: ws.cell(row=r, column=c).value for f, c in fields.items()}
        row = {"unit": label,
               "resident_code": rec.get("resident_code"),
               "resident_name": rec.get("resident_name"),
               "status": rec.get("status")}
        for k in ("total_charges", "future_charges", "prepayments", "total_owed") + AGING:
            row[k] = _num(rec.get(k)) or 0.0
        row["residential"] = not bool(NONRES.match(row["unit"]))
        cur["residents"].append(row)
    return sections, grand_row


def parse(path, strict=True):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _pick_sheet(wb)
    fields, header_row = header_map(ws, COLUMNS)

    sections, grand_row = _read_sections(ws, fields, header_row)
    residents = [r for s in sections for r in s["residents"]]
    for s in sections:
        s["summary"] = _summarise(s["residents"])
    summary = _summarise(residents)
    stopped = (grand_row["row"], "grand total") if grand_row else (
        (sections[-1]["total_row"], "section total")
        if sections and sections[-1]["total_row"] else None)

    checks, problems = [], []

    # each section has to tie to its own Total row, which is what catches a
    # section being read into the wrong property
    for s in sections:
        if s["reported_total"] is None:
            continue
        d = abs(s["summary"]["total_all"] - s["reported_total"])
        ok = d < 0.02
        checks.append({"check": f"section {s['code'] or '?'} ties to its Total row",
                       "ok": ok, "parsed": s["summary"]["total_all"],
                       "report": s["reported_total"], "delta": round(d, 2)})
        if not ok:
            problems.append(
                f"section {s['code'] or '?'} parsed {s['summary']['total_all']:,.2f} "
                f"vs its Total row (row {s['total_row']}) {s['reported_total']:,.2f}")

    gross = summary["gross_owed"]
    ag_sum = round(sum(summary["aging"].values()), 2)
    ok = abs(ag_sum - gross) < 0.02
    checks.append({"check": "aging buckets sum to gross owed", "ok": ok,
                   "aging_sum": ag_sum, "gross": gross})
    if not ok:
        problems.append(f"aging buckets sum to {ag_sum:,.2f} but gross owed is {gross:,.2f}")

    gt = grand_row or _grand_total(ws, fields)
    if gt and gt.get("total_owed") is not None:
        # the Grand Total spans every section, so compare it with all of them
        delta = abs(summary["total_all"] - gt["total_owed"])
        ok = delta < 0.02
        checks.append({"check": "total owed ties to the report Grand Total", "ok": ok,
                       "parsed": summary["total_all"], "report": gt["total_owed"],
                       "delta": round(delta, 2), "sections": len(sections)})
        if not ok:
            problems.append(
                f"parsed total owed {summary['total_all']:,.2f} across "
                f"{len(sections)} section(s) vs the report's Grand "
                f"Total (row {gt['row']}) {gt['total_owed']:,.2f}")
    else:
        checks.append({"check": "report Grand Total found", "ok": False,
                       "note": "no Grand Total row located — cannot tie out"})
        problems.append("no Grand Total row found in the export to tie out against")

    as_of = _as_of(ws)
    checks.append({"check": "as-of date found", "ok": as_of is not None,
                   "note": as_of or "no as-of date in the export — the caller should "
                                   "fall back to the file's date"})
    checks.append({"check": "stopped at a total row", "ok": stopped is not None,
                   "note": f"stopped at row {stopped[0]} ({stopped[1]!r})" if stopped
                           else "ran to the end of the sheet without a total row"})
    if stopped is None:
        problems.append("no total row after the resident rows — a summary row may "
                        "have been read as a resident")

    if not residents:
        problems.append("no resident rows found")

    if problems and strict:
        raise ValueError("delinquency report did not tie out:\n  - "
                         + "\n  - ".join(problems))

    # A single-property file keeps the shape it always had. With several
    # sections, property/property_code stay None so no caller can silently file
    # a multi-property total under whichever one happened to come first —
    # property_codes and sections carry the detail.
    if len(sections) == 1 and sections[0]["code"]:
        name, code = sections[0]["property"], sections[0]["code"]
    else:
        name, code = _property(ws, fields["unit"]) if len(sections) <= 1 else (None, None)
    return {
        "report_type": "delinquency",
        "property": name,
        "property_code": code,
        "property_codes": [s["code"] for s in sections if s["code"]],
        "as_of": as_of,
        "source_file": os.path.basename(path),
        "sheet": ws.title,
        "header_row": header_row,
        "columns": {k: v for k, v in sorted(fields.items(), key=lambda kv: kv[1])},
        "residents": residents,
        "summary": summary,
        # per-property breakdown; one entry for a single-property export
        "sections": [{"property": s["property"], "property_code": s["code"],
                      "summary": s["summary"], "rows": len(s["residents"])}
                     for s in sections],
        "checks": checks,
        "problems": problems,
    }


if __name__ == "__main__":
    out = parse(sys.argv[1])
    slim = dict(out)
    slim["residents"] = out["residents"][:3] + [f"... {len(out['residents']) - 3} more"]
    print(json.dumps(slim, indent=2, default=str))
