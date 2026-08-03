#!/usr/bin/env python3
"""Fail if personal data has reached anything that gets published or committed.

The rent roll and delinquency reports carry tenant names. The pipeline reads them
and is supposed to keep only scrubbed aggregates. This is the check that the
supposition holds — run it in CI so a regression fails the build rather than
quietly shipping names to a public URL.

Three passes:

  1. Structural — any dict key that names a person (resident_name, tenant, …)
     anywhere in the published JSON. Catches a new parser field or a renamed key.
  2. Tracked-file — any raw report (.xlsx/.csv) or per-unit pipeline output
     staged or committed to git. Those files hold everything.
  3. Value-based — if a source report is available, every surname in it is
     searched for, with word boundaries, in every published file. This is the
     one that catches a name arriving through a path nobody thought about.

Usage:
  python scripts/check_no_pii.py                          # passes 1 and 2
  python scripts/check_no_pii.py --source <report.xlsx>   # adds pass 3
"""
import argparse
import json
import os
import re
import subprocess
import sys

# Keys that identify a person. Kept in sync with build_metrics.PII_FIELDS —
# imported rather than duplicated so the two cannot drift.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from build_metrics import PII_FIELDS
except Exception:                                    # noqa: BLE001
    PII_FIELDS = ("resident_name", "resident_code", "resident", "tenant_name",
                  "tenant", "name")

PUBLISHED = ["docs/metrics.json", "docs/landing.json", "docs/scorecard.json"]
# "name" appears legitimately as a label in aggregate structures (a floorplan's
# name, a metric's name), so it is only a finding when the object it sits in
# looks like a person rather than a thing.
AMBIGUOUS = {"name", "tenant", "resident"}
PERSON_NEIGHBOURS = {"unit", "owed", "total_owed", "balance", "d30", "over90",
                     "actual_rent", "market_rent", "lease_expiration"}

problems = []


def walk(obj, path, findings):
    if isinstance(obj, dict):
        keys = set(obj)
        for k, v in obj.items():
            if k in PII_FIELDS:
                if k in AMBIGUOUS and not (keys & PERSON_NEIGHBOURS):
                    pass                              # a label on an aggregate
                else:
                    findings.append(f"{path}.{k} = {str(v)[:40]!r}")
            walk(v, f"{path}.{k}", findings)
    elif isinstance(obj, list):
        for i, v in enumerate(obj[:400]):
            walk(v, f"{path}[{i}]", findings)


def pass1_structural():
    print("1. structural — person-shaped keys in the published JSON")
    for f in PUBLISHED:
        if not os.path.exists(f):
            print(f"   SKIP {f} (not present)")
            continue
        findings = []
        walk(json.load(open(f)), os.path.basename(f), findings)
        if findings:
            print(f"   FAIL {f}")
            for x in findings[:10]:
                print(f"        {x}")
            problems.append(f"{f} contains person-shaped fields: {findings[:3]}")
        else:
            print(f"   PASS {f}")


def pass2_tracked():
    print("2. tracked files — raw reports or per-unit output in git")
    tracked = subprocess.run(["git", "ls-files"], capture_output=True, text=True).stdout.split()
    staged = subprocess.run(["git", "diff", "--cached", "--name-only"],
                            capture_output=True, text=True).stdout.split()
    bad_pat = re.compile(r"\.(xlsx|xlsm|csv)$|(^|/)_downloads/|"
                         r"(^|/)tests/fixtures/|/(rent_roll|delinquency)\.json$")
    for label, files in (("tracked", tracked), ("staged", staged)):
        hits = [f for f in files if bad_pat.search(f)]
        if hits:
            print(f"   FAIL {label}: {hits}")
            problems.append(f"{label} files that may carry names: {hits}")
        else:
            print(f"   PASS no {label} raw reports or per-unit output")


def harvest_names(source):
    """Every value sitting under a person-name header, on any sheet.

    Column-based rather than row-based: find the header cell, then take the whole
    column below it. Walking up from each data row instead misses the rent roll
    entirely, because a section marker sits between its header and its data — and
    a harvester that finds too few names makes the check pass for the wrong
    reason.
    """
    import openpyxl
    wb = openpyxl.load_workbook(source, data_only=True)
    # bare "Resident"/"Name" headers appear on the workbook's computed tabs
    # (Unit Gap Analysis, Delinquency, MTM Analysis), not just the raw reports
    pat = re.compile(r"resident (last )?name|^name$|tenant name|^tenant$|^resident$", re.I)
    names = set()
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 60) + 1):
            for c in range(1, ws.max_column + 1):
                h = ws.cell(row=r, column=c).value
                if not (isinstance(h, str) and pat.search(h.strip())):
                    continue
                for rr in range(r + 1, ws.max_row + 1):
                    v = ws.cell(row=rr, column=c).value
                    if isinstance(v, str):
                        v = v.strip()
                        # skip labels/markers and the anonymising asterisk
                        # must contain letters: numeric strings like "0.00"
                        # appear in name columns of summary blocks and would
                        # otherwise be "found" in every file that has numbers.
                        # Multi-section sheets put later section headers in the
                        # same column, so report vocabulary ("Offered rent") is
                        # rejected — nobody is named Rent.
                        if (2 <= len(v) <= 48 and re.search(r"[A-Za-z]{2}", v)
                                and not v.lower().startswith(("total", "future resident",
                                                              "summary", "grand total",
                                                              "current/notice"))
                                and not re.search(r"\b(rent|rate|lease|unit|market|offer|"
                                                  r"status|term|sq ?ft|balance|owed|notes?)\b",
                                                  v, re.I)):
                            names.add(v)
    return names


def pass3_values(source):
    print(f"3. value-based — names harvested from {os.path.basename(source)}")
    try:
        names = harvest_names(source)
    except ImportError:
        print("   SKIP openpyxl not available")
        return
    print(f"   {len(names)} distinct person-name values found in the source")
    if len(names) < 20:
        print("   WARN suspiciously few names harvested — the check may be weak")
        problems.append(f"only {len(names)} names harvested from {source}; "
                        "the value-based pass may not be meaningful")
    targets = PUBLISHED + ["docs/index.html", "docs/data.html"]
    targets += [os.path.join(dp, f) for dp, _, fs in os.walk("data")
                for f in fs if f.endswith(".json")]
    for f in targets:
        if not os.path.exists(f):
            continue
        txt = open(f, encoding="utf-8", errors="replace").read()
        leaks = [n for n in names
                 if re.search(r"(?<![A-Za-z])" + re.escape(n) + r"(?![A-Za-z])", txt)]
        if leaks:
            print(f"   FAIL {f}: {sorted(leaks)[:6]}{' …' if len(leaks) > 6 else ''}")
            problems.append(f"{f} contains {len(leaks)} name value(s): {sorted(leaks)[:4]}")
        else:
            print(f"   PASS {f}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", help="a source report to take real names from")
    a = ap.parse_args()

    os.chdir(subprocess.run(["git", "rev-parse", "--show-toplevel"],
                            capture_output=True, text=True).stdout.strip() or ".")
    pass1_structural()
    pass2_tracked()
    if a.source:
        pass3_values(a.source)
    else:
        print("3. value-based — SKIPPED (pass --source <report.xlsx> to enable)")

    print()
    if problems:
        print("FAIL: personal data reached a published or committed file", file=sys.stderr)
        for p in problems:
            print("  * " + p, file=sys.stderr)
        sys.exit(1)
    print("PASS: no personal data in any published or committed file")
