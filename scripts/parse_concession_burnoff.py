"""
parse_concession_burnoff.py
---------------------------
Parses the Yardi concession burn-off export (ConcessionBurnOffMM_DD_YYYY.xlsx).
One sheet ("Report1"), verified against the 2026-08-10 export via the inspect
workflow:

  r1  Concession Burn Off
  r2  For Selected Properties
  r3  As Of = 08/10/2026
  r4-6  a header split over three rows:
        Unit | Unit Type | Resident | Name | Move In Date | Lease Start Date |
        Total Recurring Concessions | Current Lease Concessions |
        Current Lease Concessions Remaining | Concession End Date |
        Lease Term | Market Rent | Lease Rent | Current Month
  r7+  one row per unit with an active concession, then a total row

The Resident/Name columns are read only to tell a data row from the total row
and are NEVER put in the output -- the row dicts simply do not carry them, in
addition to build_metrics.scrub() stripping PII centrally.

SECTIONS: "For Selected Properties" turned out to mean several property blocks
in one sheet -- the first real run's tie-out caught it (units summed -29,328
against a "total" of -1,833, which was the LAST property's subtotal, not a
grand total). The layout below the header is therefore walked as sections: a
text-only row opens a section and its text is the section's label; unit rows
accumulate under it; a money row with no unit number closes it as the
subtotal. A closing row arriving with no open units after other sections have
closed is the grand total. Every section must tie out against its own
subtotal, and the grand total (when present) against the sum of subtotals,
else the parse is refused.

ATTRIBUTION: each section's label is whatever the heading row says. If a label
resolves through config/properties.json (codes or aliases), that section
stores against its property; a label that resolves nowhere is warned with its
text in the log, so the log itself answers what the export covers. Labels are
taken from the Unit/Unit-Type columns only, never the resident-name columns.
"""

import datetime
import os
import re

import openpyxl

SHEET = "Report1"
TITLE = "Concession Burn Off"

# canonical keys, in sheet column order (C and D are the resident-name columns,
# read for row classification but never emitted)
COLS = ("unit", "unit_type", "_resident", "_name", "move_in", "lease_start",
        "recurring_concessions", "current_lease_concessions",
        "concessions_remaining", "concession_end", "lease_term",
        "market_rent", "lease_rent", "current_month")
MONEY = ("recurring_concessions", "current_lease_concessions",
         "concessions_remaining", "market_rent", "lease_rent")


def _num(v):
    """Yardi money cells arrive as numbers OR text — "1,234.56", "(1,500.00)",
    "$2,400", "-" — and the first real export failed the naive sum on exactly
    this. Parse what parses; everything else is None, never a guess."""
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return None
    s = str(v).strip().replace("$", "").replace(",", "")
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    try:
        return -float(s) if neg else float(s)
    except ValueError:
        return None


def _date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.date() if isinstance(v, datetime.datetime) else v).isoformat()
    return v


def parse(path, strict=True):
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"not a concession burn-off export: no {SHEET!r} sheet")
    ws = wb[SHEET]

    if str(ws["A1"].value or "").strip() != TITLE:
        raise ValueError(f"A1 is {ws['A1'].value!r}, expected {TITLE!r} -- layout moved")
    coverage = str(ws["A2"].value or "").strip() or None
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", str(ws["A3"].value or ""))
    as_of = f"{m.group(3)}-{m.group(1)}-{m.group(2)}" if m else None
    if strict and not as_of:
        raise ValueError(f"no 'As Of = MM/DD/YYYY' in A3 ({ws['A3'].value!r})")

    # the header spans rows 4-6; row 4 must still carry the anchor labels
    if str(ws["A4"].value or "").strip() != "Unit":
        raise ValueError(f"A4 is {ws['A4'].value!r}, expected 'Unit' -- header moved")

    def sums(rows):
        return {k: round(sum(u[k] or 0 for u in rows), 2) for k in MONEY}

    sections, checks = [], []
    label, units, grand_total = None, [], None

    def close(subtotal):
        nonlocal label, units
        sec = {"label": label, "unit_count": len(units),
               "totals": sums(units), "units": units,
               "subtotal": subtotal}
        for k in MONEY:
            want = None if subtotal is None else subtotal.get(k)
            if want is None:
                continue
            ok = abs(sec["totals"][k] - want) < 0.5
            checks.append({"section": label, "field": k,
                           "units_sum": sec["totals"][k],
                           "report_total": round(want, 2), "ok": ok})
            if strict and not ok:
                raise ValueError(
                    f"tie-out failed in section {label!r}: units sum to "
                    f"{sec['totals'][k]:,.2f} for {k} but its subtotal says "
                    f"{want:,.2f}")
        sections.append(sec)
        label, units = None, []

    for r in ws.iter_rows(min_row=7, values_only=True):
        vals = dict(zip(COLS, r))
        money = {k: _num(vals.get(k)) for k in MONEY}
        has_money = any(v is not None for v in money.values())
        if not has_money:
            # A heading row carries text in the Unit column ALONE (never the
            # resident-name columns); a unit row with no figures still has its
            # unit type beside it and is noise, not a heading.
            if vals.get("unit") is not None and vals.get("unit_type") is None:
                label = str(vals["unit"]).strip() or label
            continue
        if vals.get("unit") is None or                 str(vals.get("unit")).strip().lower().startswith("total"):
            # a money row with no unit closes the open section as its subtotal;
            # with nothing open after other sections closed, it is the grand total
            if units:
                close(money)
            elif sections:
                grand_total = money
            else:
                close(money)          # a total with no units still records
            continue
        units.append({
            "unit": vals.get("unit"),
            "unit_type": vals.get("unit_type"),
            "move_in": _date(vals.get("move_in")),
            "lease_start": _date(vals.get("lease_start")),
            "concession_end": _date(vals.get("concession_end")),
            "lease_term": vals.get("lease_term"),
            **money,
        })
    if units:
        close(None)                   # trailing units with no subtotal row

    all_units = [u for s in sections for u in s["units"]]
    totals = sums(all_units)
    if grand_total is not None:
        for k in MONEY:
            want = grand_total.get(k)
            if want is None:
                continue
            ok = abs(totals[k] - want) < 0.5
            checks.append({"section": "GRAND TOTAL", "field": k,
                           "units_sum": totals[k],
                           "report_total": round(want, 2), "ok": ok})
            if strict and not ok:
                raise ValueError(f"tie-out failed: sections sum to "
                                 f"{totals[k]:,.2f} for {k} but the grand "
                                 f"total says {want:,.2f}")

    return {
        "report_type": "concession_burnoff",
        "as_of": as_of,
        "source_file": os.path.basename(path),
        "coverage": coverage,
        "unit_count": len(all_units),
        "totals": totals,
        "checks": checks,
        # one routable section per property block; the label is the heading
        # row's own text and routes through properties.json codes/aliases
        "sections": [{"property_code": s["label"], "label": s["label"],
                      "unit_count": s["unit_count"], "totals": s["totals"],
                      "units": s["units"]} for s in sections],
        # kept for the no-sections fallback path and its log message
        "property_code": None,
        "unattributed": all(not s["label"] for s in sections),
    }


if __name__ == "__main__":
    import json
    import sys
    print(json.dumps(parse(sys.argv[1]), indent=2, default=str))
