#!/usr/bin/env python3
"""Extract config/coa_map.json from the COA mapping workbook.

The owner's "COA Mapping - Align l JPM l Rubicon" workbook maps other chart-of-
accounts trees onto the Align account tree. The Landing's T12 statement arrives
on the JPM tree (jpm_bf1, six-digit accounts), so the pipeline needs this map at
parse time to express its expenses in Align-tree groupings.

The workbook itself is gitignored like every xlsx; this script distils it into
committed config. Re-run it when the mapping workbook changes:

  python scripts/extract_coa_map.py "COA Mapping - Align l JPM l Rubicon .xlsx"

Account codes and names only — a chart of accounts carries no personal data.
"""
import datetime
import json
import os
import sys

import openpyxl

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "config", "coa_map.json")


def sheet_map(ws):
    out = {}
    for r in list(ws.iter_rows(values_only=True))[1:]:
        src, align = r[0], r[2] if len(r) > 2 else None
        if src and align:
            out[str(src).strip()] = [str(align).strip(),
                                     str(r[3] or "").strip() if len(r) > 3 else ""]
    return out


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    for need in ("JPM Accts", "Rubicon Accts"):
        if need not in wb.sheetnames:
            sys.exit(f"expected sheet {need!r} not found (have {wb.sheetnames})")
    out = {
        "source": os.path.basename(path),
        "extracted_at": datetime.datetime.now(datetime.timezone.utc)
                        .isoformat(timespec="seconds"),
        "jpm": sheet_map(wb["JPM Accts"]),
        "rubicon": sheet_map(wb["Rubicon Accts"]),
    }
    json.dump(out, open(OUT, "w"), indent=1)
    print(f"wrote {OUT}: {len(out['jpm'])} JPM and {len(out['rubicon'])} Rubicon mappings")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    main(sys.argv[1])
